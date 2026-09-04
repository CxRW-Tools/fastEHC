### Global variable(s)
# The size of concurrency snapshots in seconds; decreasing will provide more precision but increase processing time
CC_SNAPSHOT_SECONDS = 15
# Which severities to track per-project/per-team (avg/max/min)
ENTITY_SEVERITIES = ['critical', 'high', 'medium', 'low', 'info']
ENTITY_SEVERITY_FIELDS = {'critical': 'Critical', 'high': 'High', 'medium': 'Medium', 'low': 'Low', 'info': 'Info'}
# How many rows the Top-N project/team chart blocks hold
TOP_N_ENTITIES = 15


### Required for core functionality
import argparse
import os
import re
import ijson
from datetime import datetime, timedelta
from dateutil.parser import parse as parse_date
from collections import defaultdict
import math
import statistics
import bisect
import csv

try:
    from tqdm import tqdm
    tqdm_available = True
except ImportError:
    tqdm_available = False
    print("Consider installing tqdm for progress bar: 'pip install tqdm'")

### For direct integration with Excel workbook -- the workbook is generated entirely
### from code (see workbook_builder.py / cx_theme.py), no external template file needed
try:
    from openpyxl.utils import column_index_from_string, get_column_letter
    import workbook_builder
    import cx_theme
    xl_available = True
except ImportError:
    xl_available = False

### For debugging only
import pprint


### Wraps a binary file object so that ijson's internal .read() calls advance a tqdm byte progress bar
class _ProgressFileReader:
    def __init__(self, fileobj, pbar):
        self._fileobj = fileobj
        self._pbar = pbar

    def read(self, size=-1):
        chunk = self._fileobj.read(size)
        self._pbar.update(len(chunk))
        return chunk


### Ingest the data file
def ingest_file(file_path):
    file_size = os.path.getsize(file_path)
    scans = []
    tmp_field_names = []
    with open(file_path, 'rb') as file:
        # Extract field names from the @odata.context string (a small read near the start of the file)
        context = ijson.items(file, '@odata.context')
        context_str = next(context)
        pattern = r"#Scans\((.*?)\)"
        match = re.search(pattern, context_str)
        if match:
            fields_str = match.group(1)
            tmp_field_names = [field.strip() for field in fields_str.split(',')]
            # Adjust the field names here, using tmp_field_names
            field_names = [field.replace('(LanguageName', '') if 'ScannedLanguages' in field else field for field in tmp_field_names]

        # Reset file pointer and extract scan items, showing progress by bytes read since
        # this is a full pass over the (potentially very large) file
        file.seek(0)
        if tqdm_available:
            with tqdm(total=file_size, unit='B', unit_scale=True, unit_divisor=1024, desc="Reading data file") as pbar:
                reader = _ProgressFileReader(file, pbar)
                for scan in ijson.items(reader, 'value.item'):
                    scans.append(scan)
        else:
            print("Reading data file...", end="", flush=True)
            for scan in ijson.items(file, 'value.item'):
                scans.append(scan)
            print("completed!")

    return field_names, scans


### Calcuate the difference between two timestamps in seconds
def calculate_time_difference(t1, t2):
    dt1 = parse_date(t1)
    dt2 = parse_date(t2)
    time_diff = (dt2 - dt1).total_seconds()
    return time_diff


### Convert time in seconds to hours, minutes, and seconds
def format_seconds_to_hms(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"


### Convert time in seconds to a timeobject
def format_seconds_to_timedelta(seconds):
    return timedelta(seconds=seconds)


### Normalize a raw ProjectName into a "logical project" by stripping the leading
# team-name token (redundant with TeamName) and any branch/suffix tokens after the
# next one. Falls back to the raw name whenever the pattern doesn't apply, rather
# than guessing.
def normalize_project_name(project_name, team_name):
    if not project_name:
        return project_name or ""
    parts = project_name.split("_", 2)
    if len(parts) >= 2 and team_name and parts[0].strip().lower() == team_name.strip().lower():
        return f"{team_name}/{parts[1]}"
    return project_name


### Create a fresh stats accumulator for one project or team entity
def _new_entity_stats(team_name):
    stats = {
        'team_name': team_name,
        'COUNT_scans': 0, 'COUNT_full_scans': 0, 'COUNT_incremental_scans': 0,
        'SUM_loc': 0, 'MAX_loc': 0,
        'SUM_failed_loc': 0, 'MAX_failed_loc': 0,
        'SUM_file_count': 0, 'MAX_file_count': 0,
        'first_scan_date': None, 'last_scan_date': None,
        'active_days': set(),
        'unique_projects': set(),
    }
    for sev in ENTITY_SEVERITIES:
        stats[f'SUM_{sev}'] = 0
        stats[f'MAX_{sev}'] = 0
        stats[f'MIN_{sev}'] = None
    return stats


### Roll one scan's metrics into a project or team stats accumulator
def _accumulate_entity_stats(stats, scan, loc, scan_date):
    stats['COUNT_scans'] += 1
    if scan.get('IsIncremental'):
        stats['COUNT_incremental_scans'] += 1
    else:
        stats['COUNT_full_scans'] += 1

    stats['SUM_loc'] += loc
    stats['MAX_loc'] = max(stats['MAX_loc'], loc)

    failed_loc = scan.get('FailedLOC', 0) or 0
    stats['SUM_failed_loc'] += failed_loc
    stats['MAX_failed_loc'] = max(stats['MAX_failed_loc'], failed_loc)

    file_count = scan.get('FileCount', 0) or 0
    stats['SUM_file_count'] += file_count
    stats['MAX_file_count'] = max(stats['MAX_file_count'], file_count)

    if stats['first_scan_date'] is None or scan_date < stats['first_scan_date']:
        stats['first_scan_date'] = scan_date
    if stats['last_scan_date'] is None or scan_date > stats['last_scan_date']:
        stats['last_scan_date'] = scan_date
    stats['active_days'].add(scan_date)

    for sev in ENTITY_SEVERITIES:
        v = scan.get(ENTITY_SEVERITY_FIELDS[sev], 0) or 0
        stats[f'SUM_{sev}'] += v
        stats[f'MAX_{sev}'] = max(stats[f'MAX_{sev}'], v)
        stats[f'MIN_{sev}'] = v if stats[f'MIN_{sev}'] is None else min(stats[f'MIN_{sev}'], v)


### Compute the derived (average/percentage) fields for one project or team entity
def _finalize_entity_stats(stats, overall_total_weeks):
    count = stats['COUNT_scans']
    stats['AVG_loc'] = math.ceil(stats['SUM_loc'] / count) if count else 0
    stats['AVG_file_count'] = math.ceil(stats['SUM_file_count'] / count) if count else 0
    stats['PCT_incremental'] = stats['COUNT_incremental_scans'] / count if count else 0
    stats['active_day_count'] = len(stats['active_days'])
    stats['AVG_scans_per_week'] = stats['COUNT_scans'] / overall_total_weeks if overall_total_weeks else 0
    for sev in ENTITY_SEVERITIES:
        stats[f'AVG_{sev}'] = round(stats[f'SUM_{sev}'] / count, 2) if count else 0


### Output a data structure to the Excel workbook starting at the indicated cell (e.g., J4)
# `formats` is optional: either a flat list of one number_format string (or None) per
# column, applied to every row, or a list of lists matching `data`'s shape exactly for
# rows whose columns have mixed types (e.g. a description column followed by a date on
# one row and a plain count on the next).
def write_to_excel(ws, data, start_col, start_row, formats=None):
    try:
        # Convert start_col from letters to a numerical index
        start_col_index = column_index_from_string(start_col)
        per_row_formats = bool(formats) and isinstance(formats[0], (list, tuple))

        for row_offset, row_data in enumerate(data, start=0):
            row_formats = formats[row_offset] if per_row_formats else formats
            is_alt_row = row_offset % 2 == 1
            for col_offset, value in enumerate(row_data, start=0):
                # Calculate actual row and column indices
                row_idx = start_row + row_offset
                col_idx = start_col_index + col_offset
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cx_theme.style_body_cell(cell, alt=is_alt_row)
                if row_formats and col_offset < len(row_formats) and row_formats[col_offset]:
                    cell.number_format = row_formats[col_offset]
    except IOError as e:
        print(f"IOError when writing to Excel file: {e}")
        return
    except Exception as e:
        print(f"Unexpected error when writing to the Excel file: {e}")
        return


### Output a data structure to a csv file
def write_to_csv(header, data, filename):
    try:
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(header)
            writer.writerows(data)
    except IOError as e:
        print(f"IOError when writing to file: {e}")
    except Exception as e:
        print(f"Unexpected error when creating/writing to the CSV file: {e}")


### Write a single complete scan record to the full data csv
def write_scan_to_full_csv(field_names, scan, writer):
    try:
        # Build a row by extracting each field from the scan in the order of field_names
        row = []
        for field in field_names:
            if field == 'ScannedLanguages':
                # Special handling for ScannedLanguages field to convert list of dicts to comma-separated string
                languages = scan.get(field, [])
                language_str = ', '.join(lang['LanguageName'] for lang in languages)
                row.append(language_str)
            else:
                # For all other fields, use the value as-is
                row.append(scan.get(field, ""))
        # Write the constructed row to the CSV file
        writer.writerow(row)
    except IOError as e:
        print(f"IOError when writing to file: {e}")
    except Exception as e:
        print(f"Unexpected error when creating/writing to the CSV file: {e}")


### Process the scan data.
# One single function will be more efficient but start to get messy. Brace youreself.
def process_scans(scans, full_csv):

    ### Define (most) variables and data structures

    # Aggregate Metrics: Store high level metrics such as sums, averages, maximums, and totals
    # Note that sum, avg, and max are always associated with more granual metrics from other structures
    aggregate_metrics = {
        'COUNT_yes_scans': 0, # number of scans that fully ran
        'COUNT_no_scans': 0, # number of no-scans due to no code change
        'COUNT_missing_scans': 0, # number of scans with no recorded LOC
        'COUNT_scans': 0, # total of yes and no scans; excludes missing scans because those scans are not included in other metrics
        'COUNT_full_scans': 0, # number of full scans requested (includes no_scans)
        'COUNT_incremental_scans': 0, # number of incremental scans requested (includes no_scans)

        'SUM_loc': 0, # sum of the lines of code scanned in the data set
        'SUM_failed_loc': 0, # sum of the filed lines of code scanned in the data set
        'AVG_loc_scan': 0, # average number of lines of code per scan
        'AVG_failed_loc_scan': 0, # average number of failed lines of code per scan
        'AVG_loc_day': 0, # average number of lines of code per day
        'MAX_loc_scan': 0, # maximum number of lines of code per scan
        'MAX_failed_loc_scan': 0, # maximum number of failed lines of code per scan
        'MAX_loc_day': 0, # maximum number of lines of code per day

        'SUM_total_results': 0, # sum of total scan results
        'SUM_critical_results': 0, # sum of critical scan results
        'SUM_high_results': 0, # sum of high scan results
        'SUM_medium_results': 0, # sum of medium scan results
        'SUM_low_results': 0, # sum of low scan results
        'SUM_info_results': 0, # sum of info scan results
        'AVG_total_results': 0, # average number of total scan results
        'AVG_critical_results': 0, # average number of critical scan results
        'AVG_high_results': 0, # average number of high scan results
        'AVG_medium_results': 0, # average number of medium scan results
        'AVG_low_results': 0, # average number of low scan results
        'AVG_info_results': 0, # average number of info scan results
        'MAX_total_results': 0, # maximum number of total scan results
        'MAX_critical_results': 0, # maximum number of high scan results
        'MAX_high_results': 0, # maximum number of high scan results
        'MAX_medium_results': 0, # maximum number of medium scan results
        'MAX_low_results': 0, # maximum number of low scan results
        'MAX_info_results': 0, # maximum number of info scan results

        'COUNT_critical_results_scans': 0, # count of scans with high results
        'COUNT_high_results_scans': 0, # count of scans with high results
        'COUNT_medium_results_scans': 0, # count of scans with high results
        'COUNT_low_results_scans': 0, # count of scans with high results
        'COUNT_info_results_scans': 0, # count of scans with high results
        'COUNT_zero_results_scans': 0, # count of scans with high results

        'SUM_source_pulling_time': 0, # sum of total source pulling time in seconds
        'SUM_queue_time': 0, # sum of total queue time in seconds
        'SUM_engine_scan_time': 0, # sum of total engine scan time in seconds
        'SUM_total_scan_time': 0, # sum of total total scan time in seconds
        'AVG_source_pulling_time': 0, # average source pulling time in seconds
        'AVG_queue_time': 0, # average queue time in seconds
        'AVG_engine_scan_time': 0, # average engine scan time in seconds
        'AVG_total_scan_time': 0, # average total scan time in seconds; note this is not a sum of other times but specific data field that may exceed the sum
        'MAX_source_pulling_time': 0, # maximum source pulling time in seconds
        'MAX_queue_time': 0, # maximum queue time in seconds
        'MAX_engine_scan_time': 0, # maximum engine scan time in seconds
        'MAX_total_scan_time': 0, # maximum total scan time in seconds

        'COUNT_mon_scans': 0, # count of scans occurring on a Monday
        'COUNT_tue_scans': 0, # count of scans occurring on a Tuesday
        'COUNT_wed_scans': 0, # count of scans occurring on a Wednedsay
        'COUNT_thu_scans': 0, # count of scans occurring on a Thurdsay
        'COUNT_fri_scans': 0, # count of scans occurring on a Friday
        'COUNT_sat_scans': 0, # count of scans occurring on a Saturday
        'COUNT_sun_scans': 0, # count of scans occurring on a Sunday
        'COUNT_weekday_scans': 0, # count of scans occurring on a weekday
        'COUNT_weekend_scans': 0, # count of scans occurring on a weekend
        'MAX_scans_day': 0, # maximum number of scans per day
        'MAX_scan_date': None, # the date with the most scans

        'COUNT_projects_scanned': 0, # count of unique projects scanned

        'first_scan_date': datetime.max.date(), # the date of the first scan in the data set
        'last_scan_date': datetime.min.date(), # the date of the last scan in the data set
        'total_days': 0, # the number of days between the first and last scan
        'total_weeks': 0, # the number of weeks between the first and last scan
        'total_scan_days': 0 # the totaly number of days that actually had scans
        }

    # Languages: Store language metrics (language_name, scan_count, scan_percentage) on a dynamic list of languages
    scan_languages = {}

    # Scan Origins: Store origin metrics; this is complicated because of many custom-named origins that need to be grouped
    scan_origins = {
        'ADO': {'printable_name': 'Azure DevOps', 'scan_count': 0, 'scan_percentage': 0},
        'Bamboo': {'printable_name': 'Bamboo', 'scan_count': 0, 'scan_percentage': 0},
        'CLI': {'printable_name': 'CLI', 'scan_count': 0, 'scan_percentage': 0},
        'cx-CLI': {'printable_name': 'CxCLI', 'scan_count': 0, 'scan_percentage': 0},
        'CxFlow': {'printable_name': 'CxFlow', 'scan_count': 0, 'scan_percentage': 0},
        'Eclipse': {'printable_name': 'Eclipse', 'scan_count': 0, 'scan_percentage': 0},
        'cx-intellij': {'printable_name': 'IntelliJ', 'scan_count': 0, 'scan_percentage': 0},
        'Jenkins': {'printable_name': 'Jenkins', 'scan_count': 0, 'scan_percentage': 0},
        'Manual': {'printable_name': 'Manual', 'scan_count': 0, 'scan_percentage': 0},
        'Maven': {'printable_name': 'Maven', 'scan_count': 0, 'scan_percentage': 0},
        'Other': {'printable_name': 'Other', 'scan_count': 0, 'scan_percentage': 0},
        'System': {'printable_name': 'System', 'scan_count': 0, 'scan_percentage': 0},
        'TeamCity': {'printable_name': 'TeamCIty', 'scan_count': 0, 'scan_percentage': 0},
        'TFS': {'printable_name': 'TFS', 'scan_count': 0, 'scan_percentage': 0},
        'Visual Studio': {'printable_name': 'Visual Studio', 'scan_count': 0, 'scan_percentage': 0},
        'Visual-Studio-Code': {'printable_name': 'VS Code', 'scan_count': 0, 'scan_percentage': 0},
        'VSTS': {'printable_name': 'VSTS', 'scan_count': 0, 'scan_percentage': 0},
        'Web Portal': {'printable_name': 'Web Portal', 'scan_count': 0, 'scan_percentage': 0}
    }

    # Scan Presets: Store preset metrics (preset_name, scan_count, scan_percentage) on a dynamic list of presets
    scan_presets = {}

    # Scan Times by LOC: Store various times for every scan grouped by LOC (source_pulling_time, queue_time, engine_scan_time, total_scan_time)
    scan_times_by_loc = {
        '0-20k': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '20k-50k': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '50k-100k': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '100k-250k': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '250k-500k': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '500k-1M': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '1M-2M': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '2M-3M': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '3M-5M': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '5M-7M': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '7M-10M': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0},
        '10M+': {'COUNT_yes_scans': 0, 'COUNT_no_scans': 0, 'SUM_total_scan_time': 0, 'SUM_source_pulling_time': 0, 'SUM_queue_time': 0,
        'SUM_engine_scan_time': 0, 'MAX_total_scan_time': 0, 'MAX_source_pulling_time': 0, 'MAX_queue_time': 0, 'MAX_engine_scan_time': 0,
        'AVG_total_scan_time': 0, 'AVG_source_pulling_time': 0, 'AVG_queue_time': 0, 'AVG_engine_scan_time': 0}
    }

    # Scan Statistics by Date: Store various statistics for every scan grouped by scan date
    scan_stats_by_date = {}

    # Project & Team Stats: size/volume/severity metrics grouped by logical project and by team
    project_stats = {}
    team_stats = {}
    team_name_to_owning_ids = defaultdict(set)

    # Temporary structure to track unique projects
    temp_pids = set()

    # Variables for concurrency
    # Event format: (timestamp, change_in_count, event_type)
    # Snapshot format: (timestamp, active_engines, queue_length)
    # change_in_count is +1 for starts (entering queue or starting engine) and -1 for ends (leaving queue or engine finishing)
    # event_type distinguishes between 'queue' and 'engine'
    cc_events = filtered_cc_events = snapshot_metrics = []

    ### Prepare to output CSV of all scan data and create output file, if required
    if full_csv['enabled']:
        try:
            filename = os.path.join(full_csv['output_dir'], f'00-full_scan_data.csv')
            full_csv_file = open(filename, mode='w', newline='', encoding='utf-8')
            full_csv_writer = csv.writer(full_csv_file)
            full_csv_writer.writerow(full_csv['field_names'])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    ### Initialize tqdm object; we exclude concurrency processing because it's so fast, even for massive data sets
    if tqdm_available:
        pbar = tqdm(total=len(scans), desc="Processing scans")
    else:
        print("Processing scans...", end="", flush=True)

    ### Scan processing loop
    for scan in scans:
        if tqdm_available:
            pbar.update(1)
            pbar.refresh()

        # If required, we want to output to the full scan CSV first so as to include scans with missing fields (such as loc). This will cause a potential
        # mismatch between record counts but shouldn't impact anything relating to metrics or analysis. This CSV is only used for manual analysis.
        if full_csv['enabled']:
            write_scan_to_full_csv(full_csv['field_names'], scan, full_csv_writer)

        # If there is no LOC value, we might as well just completely skip the scan.
        # This differs from the current process but ensures that scan counts actually match in various metrics.
        # We will record the missing scan.
        loc = scan.get('LOC', None)
        if loc is None:
            aggregate_metrics['COUNT_missing_scans'] += 1
            continue
        else:
            aggregate_metrics['COUNT_scans'] += 1

        scan_date_str = scan.get('ScanRequestedOn', '').split('T')[0]
        scan_date = datetime.strptime(scan_date_str, '%Y-%m-%d').date()

        ### Populate and update key aggregate metrics; note that averages have to be calculated later and many metrics are addressed later
        if scan.get('EngineFinishedOn', None) is not None:
            noscan = False
            aggregate_metrics['COUNT_yes_scans'] += 1
        else:
            noscan = True
            aggregate_metrics['COUNT_no_scans'] += 1

        if scan.get('IsIncremental', None):
            aggregate_metrics['COUNT_incremental_scans'] += 1
        else:
            aggregate_metrics['COUNT_full_scans'] += 1

        aggregate_metrics['SUM_loc'] += loc
        aggregate_metrics['SUM_failed_loc'] += scan.get('FailedLOC', 0)
        aggregate_metrics['MAX_loc_scan'] = max(aggregate_metrics['MAX_loc_scan'], loc)
        aggregate_metrics['MAX_failed_loc_scan'] = max(aggregate_metrics['MAX_failed_loc_scan'], scan.get('FailedLOC', 0))

        aggregate_metrics['SUM_total_results'] += scan.get('TotalVulnerabilities', 0)
        aggregate_metrics['SUM_critical_results'] += scan.get('Critical', 0)
        aggregate_metrics['SUM_high_results'] += scan.get('High', 0)
        aggregate_metrics['SUM_medium_results'] += scan.get('Medium', 0)
        aggregate_metrics['SUM_low_results'] += scan.get('Low', 0)
        aggregate_metrics['SUM_info_results'] += scan.get('Info', 0)
        aggregate_metrics['MAX_total_results'] = max(aggregate_metrics['MAX_total_results'], scan.get('TotalVulnerabilities', 0))
        aggregate_metrics['MAX_critical_results'] = max(aggregate_metrics['MAX_critical_results'], scan.get('Critical', 0))
        aggregate_metrics['MAX_high_results'] = max(aggregate_metrics['MAX_high_results'], scan.get('High', 0))
        aggregate_metrics['MAX_medium_results'] = max(aggregate_metrics['MAX_medium_results'], scan.get('Medium', 0))
        aggregate_metrics['MAX_low_results'] = max(aggregate_metrics['MAX_low_results'], scan.get('Low', 0))
        aggregate_metrics['MAX_info_results'] = max(aggregate_metrics['MAX_info_results'], scan.get('Info', 0))
        if scan.get('Critical', 0) > 0:
            aggregate_metrics['COUNT_critical_results_scans'] += 1
        if scan.get('High', 0) > 0:
            aggregate_metrics['COUNT_high_results_scans'] += 1
        if scan.get('Medium', 0) > 0:
            aggregate_metrics['COUNT_medium_results_scans'] += 1
        if scan.get('Low', 0) > 0:
            aggregate_metrics['COUNT_low_results_scans'] += 1
        if scan.get('Info', 0) > 0:
            aggregate_metrics['COUNT_info_results_scans'] += 1
        if scan.get('TotalVulnerabilities', 0) == 0:
            aggregate_metrics['COUNT_zero_results_scans'] += 1

        # Update time metrics; deal with negative duration issues that sometimes occur
        source_pulling_time = max(calculate_time_difference(scan.get('ScanRequestedOn'),scan.get('QueuedOn')), 0)
        aggregate_metrics['SUM_source_pulling_time'] += source_pulling_time
        aggregate_metrics['MAX_source_pulling_time'] = max(aggregate_metrics['MAX_source_pulling_time'], source_pulling_time)

        queue_time = max(calculate_time_difference(scan.get('QueuedOn'),scan.get('EngineStartedOn')), 0)
        aggregate_metrics['SUM_queue_time'] += queue_time
        aggregate_metrics['MAX_queue_time'] = max(aggregate_metrics['MAX_queue_time'], queue_time)

        if noscan is False:
            engine_scan_time = max(calculate_time_difference(scan.get('EngineStartedOn'),scan.get('EngineFinishedOn')), 0)
            aggregate_metrics['SUM_engine_scan_time'] += engine_scan_time
            aggregate_metrics['MAX_engine_scan_time'] = max(aggregate_metrics['MAX_engine_scan_time'], calculate_time_difference(scan.get('EngineStartedOn'),scan.get('EngineFinishedOn')))

        total_scan_time = max(calculate_time_difference(scan.get('ScanRequestedOn'),scan.get('ScanCompletedOn')), 0)
        aggregate_metrics['SUM_total_scan_time'] += total_scan_time
        aggregate_metrics['MAX_total_scan_time'] = max(aggregate_metrics['MAX_total_scan_time'], total_scan_time)

        # Increment the proper day counters
        day_of_week = scan_date.strftime('%A')
        day_key_map = {
            'Monday': 'COUNT_mon_scans',
            'Tuesday': 'COUNT_tue_scans',
            'Wednesday': 'COUNT_wed_scans',
            'Thursday': 'COUNT_thu_scans',
            'Friday': 'COUNT_fri_scans',
            'Saturday': 'COUNT_sat_scans',
            'Sunday': 'COUNT_sun_scans',
        }
        if day_of_week in day_key_map:
            day_key = day_key_map[day_of_week]
            aggregate_metrics[day_key] += 1
        if day_of_week in ['Saturday', 'Sunday']:
            aggregate_metrics['COUNT_weekend_scans'] += 1
        else:
            aggregate_metrics['COUNT_weekday_scans'] += 1

        aggregate_metrics['first_scan_date'] = min(aggregate_metrics['first_scan_date'], scan_date)
        aggregate_metrics['last_scan_date'] = max(aggregate_metrics['last_scan_date'], scan_date)

        # Increment unique project count; sometimes the Id or Name is empty so create a new key type
        project_id = scan.get('ProjectId', 0)
        project_name = scan.get('ProjectName', '')
        pid = str(project_id) + '_' + project_name
        if pid not in temp_pids:
            temp_pids.add(pid)
            aggregate_metrics['COUNT_projects_scanned'] += 1

        ### Add scanned languages
        for language in scan.get('ScannedLanguages', []):
            lang_name = language.get('LanguageName')
            if lang_name and lang_name != 'Common':
                scan_languages[lang_name] = scan_languages.get(lang_name, 0) + 1

        ### Add scan origin
        origin_key = scan.get('Origin', 'Other')
        group = next((key for key in scan_origins if origin_key.startswith(key)), 'Other')
        scan_origins[group]['scan_count'] += 1

        ### Add scan preset
        preset_name = scan.get('PresetName')
        scan_presets[preset_name] = scan_presets.get(preset_name, 0) + 1

        ### Roll this scan into its logical project and team stats
        team_name = scan.get('TeamName') or 'Unknown'
        logical_project = normalize_project_name(project_name, team_name)

        if logical_project not in project_stats:
            project_stats[logical_project] = _new_entity_stats(team_name)
        pstat = project_stats[logical_project]
        pstat['team_name'] = team_name  # last-observed team wins
        _accumulate_entity_stats(pstat, scan, loc, scan_date)

        team_name_to_owning_ids[team_name].add(scan.get('OwningTeamId'))
        if team_name not in team_stats:
            team_stats[team_name] = _new_entity_stats(None)
        tstat = team_stats[team_name]
        _accumulate_entity_stats(tstat, scan, loc, scan_date)
        tstat['unique_projects'].add(logical_project)

        ### Add scan times
        if loc <= 20000:
            bin_key = '0-20k'
        elif loc <= 50000:
            bin_key = '20k-50k'
        elif loc <= 100000:
            bin_key = '50k-100k'
        elif loc <= 250000:
            bin_key = '100k-250k'
        elif loc <= 500000:
            bin_key = '250k-500k'
        elif loc <= 1000000:
            bin_key = '500k-1M'
        elif loc <= 2000000:
            bin_key = '1M-2M'
        elif loc <= 3000000:
            bin_key = '2M-3M'
        elif loc <= 5000000:
            bin_key = '3M-5M'
        elif loc <= 7000000:
            bin_key = '5M-7M'
        elif loc <= 10000000:
            bin_key = '7M-10M'
        else:
            bin_key = '10M+'

        bin = scan_times_by_loc[bin_key]

        bin['SUM_source_pulling_time'] += source_pulling_time
        bin['SUM_queue_time'] += queue_time
        bin['SUM_total_scan_time'] += total_scan_time
        bin['MAX_source_pulling_time'] = max(source_pulling_time, bin['MAX_source_pulling_time'])
        bin['MAX_queue_time'] = max(queue_time, bin['MAX_queue_time'])
        bin['MAX_total_scan_time'] = max(total_scan_time, bin['MAX_total_scan_time'])

        if noscan is False:
            bin['COUNT_yes_scans'] += 1
            bin['SUM_engine_scan_time'] += engine_scan_time
            bin['MAX_engine_scan_time'] = max(engine_scan_time, bin['MAX_engine_scan_time'])
        else:
            bin['COUNT_no_scans'] += 1

        ### Populate scan statistics by date
        if scan_date not in scan_stats_by_date:
            scan_stats_by_date[scan_date] = {
                'COUNT_yes_scans': 0,
                'COUNT_no_scans': 0,
                'COUNT_scans': 0,
                'COUNT_full_scans': 0,
                'COUNT_incremental_scans': 0,
                'SUM_loc': 0,
                'MAX_loc': 0,
                'SUM_failed_loc': 0,
                'MAX_failed_loc': 0,
                'SUM_total_scan_time': 0,
                'SUM_source_pulling_time': 0,
                'SUM_queue_time': 0,
                'SUM_engine_scan_time': 0,
                'MAX_total_scan_time': 0,
                'MAX_source_pulling_time': 0,
                'MAX_queue_time': 0,
                'MAX_engine_scan_time': 0,
                'AVG_total_scan_time': 0,
                'AVG_source_pulling_time': 0,
                'AVG_queue_time': 0,
                'AVG_engine_scan_time': 0
            }

        if noscan is False:
            scan_stats_by_date[scan_date]['COUNT_yes_scans'] += 1
            scan_stats_by_date[scan_date]['SUM_engine_scan_time'] += engine_scan_time
            scan_stats_by_date[scan_date]['MAX_engine_scan_time'] = max(engine_scan_time, scan_stats_by_date[scan_date]['MAX_engine_scan_time'])
        else:
            scan_stats_by_date[scan_date]['COUNT_no_scans'] += 1

        if scan.get('IsIncremental', None):
            scan_stats_by_date[scan_date]['COUNT_incremental_scans'] += 1
        else:
            scan_stats_by_date[scan_date]['COUNT_full_scans'] += 1

        scan_stats_by_date[scan_date]['COUNT_scans'] += 1
        scan_stats_by_date[scan_date]['SUM_loc'] += loc
        scan_stats_by_date[scan_date]['MAX_loc'] = max(loc, scan_stats_by_date[scan_date]['MAX_loc'])
        scan_stats_by_date[scan_date]['SUM_failed_loc'] += scan.get('FailedLOC', 0)
        scan_stats_by_date[scan_date]['MAX_failed_loc'] = max(scan.get('FailedLOC', 0), scan_stats_by_date[scan_date]['MAX_failed_loc'])
        scan_stats_by_date[scan_date]['SUM_total_scan_time'] += total_scan_time
        scan_stats_by_date[scan_date]['MAX_total_scan_time'] = max(total_scan_time, scan_stats_by_date[scan_date]['MAX_total_scan_time'])
        scan_stats_by_date[scan_date]['SUM_source_pulling_time'] += source_pulling_time
        scan_stats_by_date[scan_date]['MAX_source_pulling_time'] = max(source_pulling_time, scan_stats_by_date[scan_date]['MAX_source_pulling_time'])
        scan_stats_by_date[scan_date]['SUM_queue_time'] += queue_time
        scan_stats_by_date[scan_date]['MAX_queue_time'] = max(queue_time, scan_stats_by_date[scan_date]['MAX_queue_time'])

        # Parse timestamps for concurrency queueing and engine events
        queued_on = parse_date(scan['QueuedOn']).timestamp()
        engine_started_on = parse_date(scan['EngineStartedOn']).timestamp()
        engine_finished_on = None
        optimal_scan_finish = None

        cc_events.append((queued_on, +1, 'queue'))
        cc_events.append((engine_started_on, -1, 'queue'))

        if 'EngineFinishedOn' in scan and scan['EngineFinishedOn'] is not None:
            engine_finished_on = parse_date(scan['EngineFinishedOn']).timestamp()
            engine_scan_duration = engine_finished_on - engine_started_on
            optimal_scan_finish = queued_on + engine_scan_duration  # Calculate based on no queue delay assumption
            cc_events.append((engine_started_on, +1, 'engine'))
            cc_events.append((optimal_scan_finish, -1, 'engine'))

    # End of scan processing loop

    ### Calculate metrics that require the full data set
    # Note that some averages are calculated against all scans and some ignore noscans; this matches the prior version of the tool but could be changed
    aggregate_metrics['total_days'] = (aggregate_metrics['last_scan_date'] - aggregate_metrics['first_scan_date']).days
    aggregate_metrics['total_weeks'] = math.ceil(aggregate_metrics['total_days'] / 7)
    aggregate_metrics['total_scan_days'] = len(scan_stats_by_date)

    aggregate_metrics['AVG_loc_scan'] = math.ceil(aggregate_metrics['SUM_loc'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_failed_loc_scan'] = math.ceil(aggregate_metrics['SUM_failed_loc'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_loc_day'] = math.ceil(aggregate_metrics['SUM_loc'] / (aggregate_metrics['total_scan_days']))

    for date, stats in scan_stats_by_date.items():
        aggregate_metrics['MAX_loc_day'] = max(aggregate_metrics['MAX_loc_day'], stats['SUM_loc'])

        if stats['COUNT_scans'] > aggregate_metrics['MAX_scans_day']:
            aggregate_metrics['MAX_scans_day'] = stats['COUNT_scans']
            aggregate_metrics['MAX_scan_date'] = date

        if stats['COUNT_yes_scans'] > 0:
            stats['AVG_total_scan_time'] = stats['SUM_total_scan_time'] / stats['COUNT_yes_scans']
            stats['AVG_source_pulling_time'] = stats['SUM_source_pulling_time'] / stats['COUNT_yes_scans']
            stats['AVG_queue_time'] = stats['SUM_queue_time'] / stats['COUNT_yes_scans']
            stats['AVG_engine_scan_time'] = stats['SUM_engine_scan_time'] / stats['COUNT_yes_scans']
        else:
            stats['AVG_total_scan_time'] = 0
            stats['AVG_source_pulling_time'] = 0
            stats['AVG_queue_time'] = 0
            stats['AVG_engine_scan_time'] = 0

    for bin_key, bin in scan_times_by_loc.items():
        if bin['COUNT_yes_scans'] > 0:
            bin['AVG_source_pulling_time'] = math.ceil(bin['SUM_source_pulling_time'] / (bin['COUNT_yes_scans']))
            bin['AVG_queue_time'] = math.ceil(bin['SUM_queue_time'] / (bin['COUNT_yes_scans']))
            bin['AVG_engine_scan_time'] = math.ceil(bin['SUM_engine_scan_time'] / bin['COUNT_yes_scans'])
            bin['AVG_total_scan_time'] = math.ceil(bin['SUM_total_scan_time'] / bin['COUNT_yes_scans'])

    if aggregate_metrics['COUNT_scans'] > 0:
        aggregate_metrics['AVG_source_pulling_time'] = math.ceil(aggregate_metrics['SUM_source_pulling_time'] / aggregate_metrics['COUNT_yes_scans'])
        aggregate_metrics['AVG_queue_time'] = math.ceil(aggregate_metrics['SUM_queue_time'] / aggregate_metrics['COUNT_yes_scans'])
        aggregate_metrics['AVG_total_scan_time'] = math.ceil(aggregate_metrics['SUM_total_scan_time'] / aggregate_metrics['COUNT_yes_scans'])
    if aggregate_metrics['COUNT_yes_scans'] > 0:
        aggregate_metrics['AVG_engine_scan_time'] = math.ceil(aggregate_metrics['SUM_engine_scan_time'] / aggregate_metrics['COUNT_yes_scans'])
        aggregate_metrics['AVG_total_scan_time'] = math.ceil(aggregate_metrics['SUM_total_scan_time'] / aggregate_metrics['COUNT_yes_scans'])

    aggregate_metrics['AVG_total_results'] = math.ceil(aggregate_metrics['SUM_total_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_critical_results'] = round(aggregate_metrics['SUM_critical_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_high_results'] = round(aggregate_metrics['SUM_high_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_medium_results'] = round(aggregate_metrics['SUM_medium_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_low_results'] = round(aggregate_metrics['SUM_low_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_info_results']= round(aggregate_metrics['SUM_info_results'] / aggregate_metrics['COUNT_scans'])

    for origin, data in scan_origins.items():
        data['scan_percentage'] = data['scan_count'] / aggregate_metrics['COUNT_scans']

    # Finalize project/team derived metrics now that total_weeks is known
    for stats in project_stats.values():
        _finalize_entity_stats(stats, aggregate_metrics['total_weeks'])
    for stats in team_stats.values():
        _finalize_entity_stats(stats, aggregate_metrics['total_weeks'])
    reorged_teams = {name: ids for name, ids in team_name_to_owning_ids.items() if len(ids) > 1}

    if tqdm_available:
        pbar.close()
    else:
        print("completed!")

    if reorged_teams:
        print(f"Note: {len(reorged_teams)} team name(s) were seen under more than one OwningTeamId "
              f"(likely a reorg/rename) and were merged by name: {', '.join(sorted(reorged_teams))}")

    # Process concurrency events
    print(f"Calculating scan concurrency using {CC_SNAPSHOT_SECONDS} second snapshots...", end="", flush=True)

    # Initialize variables
    cc_window_start_ts = datetime.combine(aggregate_metrics['first_scan_date'], datetime.min.time()).timestamp()
    cc_window_end_ts = datetime.combine(aggregate_metrics['last_scan_date'], datetime.min.time()).timestamp()
    num_snapshots = math.ceil((cc_window_end_ts - cc_window_start_ts) / CC_SNAPSHOT_SECONDS)

    # Filter out objects based on the window and sort them
    filtered_cc_events = [event for event in cc_events if cc_window_start_ts <= event[0] <= cc_window_end_ts]
    filtered_cc_events.sort(key=lambda x: x[0])

    current_active_engines = 0
    current_queue_length = 0
    event_index = 0
    snapshot_metrics = []

    # For each snapshot...
    for snapshot in range(num_snapshots):
        # Calculate the bounds of the snapshot in timestamp format
        snapshot_start_ts = cc_window_start_ts + snapshot * CC_SNAPSHOT_SECONDS
        next_snapshot_start_ts = snapshot_start_ts + CC_SNAPSHOT_SECONDS

        while event_index < len(filtered_cc_events) and filtered_cc_events[event_index][0] < next_snapshot_start_ts:
            event_time, change, event_type = filtered_cc_events[event_index]

            if event_type == 'engine':
                current_active_engines += change
            elif event_type == 'queue':
                current_queue_length += change

            event_index += 1

        # Convert snapshot_start_ts to datetime for recording
        snapshot_start_dt = datetime.fromtimestamp(snapshot_start_ts)

        # Append the metrics for the current snapshot to the list
        snapshot_metrics.append((snapshot_start_dt, current_active_engines, current_queue_length))
    print("completed!")

    # Close the CSV file if it's open
    if full_csv['enabled']:
        full_csv_file.close()

    return {
        'aggregate_metrics': aggregate_metrics,
        'scan_languages': scan_languages,
        'scan_origins': scan_origins,
        'scan_presets': scan_presets,
        'scan_times_by_loc': scan_times_by_loc,
        'scan_stats_by_date': scan_stats_by_date,
        'cc_metrics': snapshot_metrics,
        'project_stats': project_stats,
        'team_stats': team_stats,
    }


### Shell function to handle outputs
def output_analysis(data, csv_config, excel_config):
    # Identify daily max concurrency values based on the granular calculations made previously
    daily_maxima = defaultdict(lambda: {'actual': 0, 'optimal': 0})
    for snapshot in data['cc_metrics']:
        snapshot_dt, active_engines, queue_length = snapshot
        snapshot_date = snapshot_dt.date()
        optimal_concurrency = active_engines + queue_length

        # Update daily maximums
        daily_record = daily_maxima[snapshot_date]
        daily_record['actual'] = max(daily_record['actual'], active_engines)
        daily_record['optimal'] = max(daily_record['optimal'], optimal_concurrency)

    output_summary_of_scans(data, csv_config, excel_config)
    output_scan_metrics(data, csv_config, excel_config)
    output_scan_duration(data, csv_config, excel_config)
    output_scan_results_and_severity(data, csv_config, excel_config)
    output_scan_languages(data, csv_config, excel_config)
    output_scan_submission_summary(data, csv_config, excel_config)
    output_day_of_week_scan_average(data, csv_config, excel_config)
    output_scan_origins(data, csv_config, excel_config)
    output_scan_presets(data, csv_config, excel_config)
    output_scan_time_analysis(data, csv_config, excel_config)
    output_scan_concurrency(daily_maxima, csv_config, excel_config)
    output_scans_by_date(data['scan_stats_by_date'], csv_config, excel_config)
    output_scans_by_week(data['scan_stats_by_date'], csv_config, excel_config)
    output_project_stats(data['project_stats'], csv_config, excel_config)
    output_team_stats(data['team_stats'], csv_config, excel_config)
    output_top_projects(data['project_stats'], csv_config, excel_config)
    output_top_teams(data['team_stats'], csv_config, excel_config)
    output_project_distribution(data['project_stats'], csv_config, excel_config)
    output_team_distribution(data['team_stats'], csv_config, excel_config)


### Output Functions: Handle all types of output for a specific metric type or report section

def output_summary_of_scans(data, csv_config, excel_config):
    submitted_scans = data['aggregate_metrics']['COUNT_scans'] + data['aggregate_metrics']['COUNT_missing_scans'];
    # Create the data structure to hold the various fields
    output_data = [
        ['Start Date',data['aggregate_metrics']['first_scan_date']],
        ['End Date',data['aggregate_metrics']['last_scan_date']],
        ['Days',data['aggregate_metrics']['total_days']],
        ['Weeks',data['aggregate_metrics']['total_weeks']],
        ['Scans Submitted',submitted_scans],
        ['Scans Completed',data['aggregate_metrics']['COUNT_scans']],
        ['Scans Failed',data['aggregate_metrics']['COUNT_missing_scans']],
        ['Full Scans Submitted',data['aggregate_metrics']['COUNT_full_scans'],data['aggregate_metrics']['COUNT_full_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Incremental Scans Submitted',data['aggregate_metrics']['COUNT_incremental_scans'],data['aggregate_metrics']['COUNT_incremental_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['No-Change Scans',data['aggregate_metrics']['COUNT_no_scans'],data['aggregate_metrics']['COUNT_no_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Scans with Critical Results',data['aggregate_metrics']['COUNT_critical_results_scans'],data['aggregate_metrics']['COUNT_critical_results_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Scans with High Results',data['aggregate_metrics']['COUNT_high_results_scans'],data['aggregate_metrics']['COUNT_high_results_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Scans with Medium Results',data['aggregate_metrics']['COUNT_medium_results_scans'],data['aggregate_metrics']['COUNT_medium_results_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Scans with Low Results',data['aggregate_metrics']['COUNT_low_results_scans'],data['aggregate_metrics']['COUNT_low_results_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Scans with Informational Results',data['aggregate_metrics']['COUNT_info_results_scans'],data['aggregate_metrics']['COUNT_info_results_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Scans with Zero Results',data['aggregate_metrics']['COUNT_zero_results_scans'],data['aggregate_metrics']['COUNT_zero_results_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Unique Projects Scanned',data['aggregate_metrics']['COUNT_projects_scanned']]
    ]

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Description','Value','%'], output_data, os.path.join(csv_config['output_dir'], f'01-summary_of_scans.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        I, P, D = cx_theme.FMT_INT, cx_theme.FMT_PCT, cx_theme.FMT_DATE
        formats = [
            [None, D], [None, D], [None, I], [None, I], [None, I],
            [None, I], [None, I],
            [None, I, P], [None, I, P], [None, I, P],
            [None, I, P], [None, I, P], [None, I, P],
            [None, I, P], [None, I, P], [None, I, P],
            [None, I],
        ]
        write_to_excel(excel_config['workbook']['Data'], output_data, 'B', 4, formats)


def output_scan_metrics(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = [
        ['LOC per Scan',data['aggregate_metrics']['AVG_loc_scan'],data['aggregate_metrics']['MAX_loc_scan']],
        ['Failed LOC per Scan',data['aggregate_metrics']['AVG_failed_loc_scan'],data['aggregate_metrics']['MAX_failed_loc_scan']],
        ['Daily LOC',data['aggregate_metrics']['AVG_loc_day'],data['aggregate_metrics']['MAX_loc_day']]
    ]

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Description','Average','Max'], output_data, os.path.join(csv_config['output_dir'], f'02-scan_metrics.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'F', 4,
                        [None, cx_theme.FMT_INT, cx_theme.FMT_INT])


def output_scan_duration(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = [
        ['Total Scan Duration',format_seconds_to_timedelta(data['aggregate_metrics']['AVG_total_scan_time']),format_seconds_to_timedelta(data['aggregate_metrics']['MAX_total_scan_time'])],
        ['Source Pulling Duration',format_seconds_to_timedelta(data['aggregate_metrics']['AVG_source_pulling_time']),format_seconds_to_timedelta(data['aggregate_metrics']['MAX_source_pulling_time'])],
        ['Queued Duration',format_seconds_to_timedelta(data['aggregate_metrics']['AVG_queue_time']),format_seconds_to_timedelta(data['aggregate_metrics']['MAX_queue_time'])],
        ['Engine Scan Duration',format_seconds_to_timedelta(data['aggregate_metrics']['AVG_engine_scan_time']),format_seconds_to_timedelta(data['aggregate_metrics']['MAX_engine_scan_time'])]
    ]

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Description','Average','Max'], output_data, os.path.join(csv_config['output_dir'], f'03-scan_duration.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'J', 4,
                        [None, cx_theme.FMT_DURATION, cx_theme.FMT_DURATION])


def output_scan_results_and_severity(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = [
        ['Total',data['aggregate_metrics']['AVG_total_results'],data['aggregate_metrics']['MAX_total_results']],
        ['Critical',data['aggregate_metrics']['AVG_critical_results'],data['aggregate_metrics']['MAX_critical_results']],
        ['High',data['aggregate_metrics']['AVG_high_results'],data['aggregate_metrics']['MAX_high_results']],
        ['Medium',data['aggregate_metrics']['AVG_medium_results'],data['aggregate_metrics']['MAX_medium_results']],
        ['Low',data['aggregate_metrics']['AVG_low_results'],data['aggregate_metrics']['MAX_low_results']],
        ['Informational',data['aggregate_metrics']['AVG_info_results'],data['aggregate_metrics']['MAX_info_results']]
    ]

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Description','Average','Max'], output_data, os.path.join(csv_config['output_dir'], f'04-scan_results_severity.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        # Only overwrite the Value column so the pre-styled severity label column (N) is left alone
        ws = excel_config['workbook']['Data']
        write_to_excel(ws, [[row[1], row[2]] for row in output_data], 'O', 4,
                        [cx_theme.FMT_INT, cx_theme.FMT_INT])


def output_scan_languages(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = []
    for language, count in data['scan_languages'].items():
        output_data.append([language, count / data['aggregate_metrics']['COUNT_scans'], count])

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Language','%','Scans'], output_data, os.path.join(csv_config['output_dir'], f'05-languages.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'R', 4,
                        [None, cx_theme.FMT_PCT, cx_theme.FMT_INT])


def output_scan_submission_summary(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = [
        ['Average Scans Submitted per Week',data['aggregate_metrics']['COUNT_scans'] / data['aggregate_metrics']['total_weeks']],
        ['Average Scans Submitted per Day',data['aggregate_metrics']['COUNT_scans'] / data['aggregate_metrics']['total_days']],
        ['Average Scans Submitted per Weekday',data['aggregate_metrics']['COUNT_weekday_scans'] / (5 * data['aggregate_metrics']['total_weeks'])],
        ['Average Scans Submitted per Weekend Day',data['aggregate_metrics']['COUNT_weekend_scans'] / (2 * data['aggregate_metrics']['total_weeks'])],
        ['Max Daily Scans Submitted',data['aggregate_metrics']['MAX_scans_day']],
        ['Date of Max Daily Scans',data['aggregate_metrics']['MAX_scan_date']]
    ]

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Description','Value'], output_data, os.path.join(csv_config['output_dir'], f'06-scan_submissison_summary.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        D2, I, D = cx_theme.FMT_DECIMAL2, cx_theme.FMT_INT, cx_theme.FMT_DATE
        formats = [[None, D2], [None, D2], [None, D2], [None, D2], [None, I], [None, D]]
        write_to_excel(excel_config['workbook']['Data'], output_data, 'V', 4, formats)

def output_day_of_week_scan_average(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = [
        ['Monday',data['aggregate_metrics']['COUNT_mon_scans'], data['aggregate_metrics']['COUNT_mon_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Tuesday',data['aggregate_metrics']['COUNT_tue_scans'], data['aggregate_metrics']['COUNT_tue_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Wednesday',data['aggregate_metrics']['COUNT_wed_scans'], data['aggregate_metrics']['COUNT_wed_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Thursday',data['aggregate_metrics']['COUNT_thu_scans'], data['aggregate_metrics']['COUNT_thu_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Friday',data['aggregate_metrics']['COUNT_fri_scans'], data['aggregate_metrics']['COUNT_fri_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Saturday',data['aggregate_metrics']['COUNT_sat_scans'], data['aggregate_metrics']['COUNT_sat_scans'] / data['aggregate_metrics']['COUNT_scans']],
        ['Sunday',data['aggregate_metrics']['COUNT_sun_scans'], data['aggregate_metrics']['COUNT_sun_scans'] / data['aggregate_metrics']['COUNT_scans']]
    ]

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Day of Week','Scans','%'], output_data, os.path.join(csv_config['output_dir'], f'07-day_of_week_scan_average.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'Y', 4,
                        [None, cx_theme.FMT_INT, cx_theme.FMT_PCT])

def output_scan_origins(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = []
    for key, value in data['scan_origins'].items():
        if value['scan_count'] > 0:
            output_data.append([value['printable_name'], value['scan_count'], value['scan_percentage']])

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Origin','Scans','%'], output_data, os.path.join(csv_config['output_dir'], f'08-origins.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'AC', 4,
                        [None, cx_theme.FMT_INT, cx_theme.FMT_PCT])

def output_scan_presets(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = []
    for preset, count in data['scan_presets'].items():
        output_data.append([preset, count, count / data['aggregate_metrics']['COUNT_scans']])

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Preset','Scans','%'], output_data, os.path.join(csv_config['output_dir'], f'09-presets.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'AG', 4,
                        [None, cx_theme.FMT_INT, cx_theme.FMT_PCT])

def output_scan_time_analysis(data, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = []
    for key, value in data['scan_times_by_loc'].items():
        output_data.append([key, value['COUNT_yes_scans'] + value['COUNT_no_scans'], (value['COUNT_yes_scans'] + value['COUNT_no_scans']) / data['aggregate_metrics']['COUNT_scans'],
            format_seconds_to_timedelta(value['AVG_total_scan_time']), format_seconds_to_timedelta(value['AVG_source_pulling_time']),
            format_seconds_to_timedelta(value['AVG_queue_time']), format_seconds_to_timedelta(value['AVG_engine_scan_time'])])

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['LOC Range','Scans','% Scans','Avg Total Time','Avg Source Pulling Time','Avg Queue Time','Avg Engine Scan Time'], output_data, os.path.join(csv_config['output_dir'], f'10-scan_time_analysis.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'AK', 4,
                        [None, cx_theme.FMT_INT, cx_theme.FMT_PCT, cx_theme.FMT_DURATION,
                         cx_theme.FMT_DURATION, cx_theme.FMT_DURATION, cx_theme.FMT_DURATION])

def output_scan_concurrency(daily_maxima, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = []
    for date, maxima in sorted(daily_maxima.items()):
        output_data.append([date, maxima['actual'], maxima['optimal']])

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Date','Max Actual','Max Optimal'], output_data, os.path.join(csv_config['output_dir'], f'11-concurrency_analysis.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'AS', 4,
                        [cx_theme.FMT_DATE, cx_theme.FMT_INT, cx_theme.FMT_INT])


def output_scans_by_date(scan_stats_by_date, csv_config, excel_config):
    # Create the data structure to hold the various fields
    output_data = []
    for date, value in scan_stats_by_date.items():
        output_data.append([date, value['COUNT_scans'], value['COUNT_no_scans'], value['COUNT_full_scans'], value['COUNT_incremental_scans'],
            value['SUM_loc'],value['MAX_loc'],value['SUM_failed_loc'],value['MAX_failed_loc'],
            format_seconds_to_timedelta(value['AVG_total_scan_time']),format_seconds_to_timedelta(value['MAX_total_scan_time']),
            format_seconds_to_timedelta(value['AVG_source_pulling_time']),format_seconds_to_timedelta(value['MAX_source_pulling_time']),
            format_seconds_to_timedelta(value['AVG_queue_time']),format_seconds_to_timedelta(value['MAX_queue_time']),
            format_seconds_to_timedelta(value['AVG_engine_scan_time']),format_seconds_to_timedelta(value['MAX_engine_scan_time'])])

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Date','Scans','No Scans','Full Scans','Incremental Scans','Sum LOC','Max LOC','Sum Failed LOC','Max Failed LOC',
            'AVG Total Scan Time','Max Total Scan Time','Avg Source Pulling Time','Max Source Pulling Time','Avg Queue Time','Max Queue Time',
            'Avg Engine Time','Max Engine Time'], output_data, os.path.join(csv_config['output_dir'], f'12-scans_by_date.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'AW', 4, _SCANS_BY_PERIOD_FORMATS)


def output_scans_by_week(scan_stats_by_date, csv_config, excel_config):
    # Convert daily data to weekly data
    weekly_data = {}

    for date, value in scan_stats_by_date.items():
        # Calculate the Monday of the current week
        monday_of_week = date - timedelta(days=date.weekday())

        # Initialize or update the weekly data
        if monday_of_week not in weekly_data:
            weekly_data[monday_of_week] = value.copy()
            weekly_data[monday_of_week]['days_counted'] = 1
        else:
            week_data = weekly_data[monday_of_week]
            for key in value:
                if 'COUNT' in key or 'SUM' in key:
                    week_data[key] += value[key]
                elif 'MAX' in key:
                    week_data[key] = max(week_data[key], value[key])
            week_data['days_counted'] += 1

    # Prepare the data for output
    output_data = []  # This will be a list of lists for CSV/Excel rows
    for monday, data in weekly_data.items():
        # Calculate averages per scan if 'COUNT_yes_scans' is greater than 0 to avoid division by zero
        avg_total_scan_time = data['SUM_total_scan_time'] / data['COUNT_yes_scans'] if data['COUNT_yes_scans'] > 0 else 0
        avg_source_pulling_time = data['SUM_source_pulling_time'] / data['COUNT_yes_scans'] if data['COUNT_yes_scans'] > 0 else 0
        avg_queue_time = data['SUM_queue_time'] / data['COUNT_yes_scans'] if data['COUNT_yes_scans'] > 0 else 0
        avg_engine_scan_time = data['SUM_engine_scan_time'] / data['COUNT_yes_scans'] if data['COUNT_yes_scans'] > 0 else 0

        row = [
            monday,
            data['COUNT_scans'],
            data['COUNT_no_scans'],
            data['COUNT_full_scans'],
            data['COUNT_incremental_scans'],
            data['SUM_loc'],
            data['MAX_loc'],
            data['SUM_failed_loc'],
            data['MAX_failed_loc'],
            format_seconds_to_timedelta(avg_total_scan_time),
            format_seconds_to_timedelta(data['MAX_total_scan_time']),
            format_seconds_to_timedelta(avg_source_pulling_time),
            format_seconds_to_timedelta(data['MAX_source_pulling_time']),
            format_seconds_to_timedelta(avg_queue_time),
            format_seconds_to_timedelta(data['MAX_queue_time']),
            format_seconds_to_timedelta(avg_engine_scan_time),
            format_seconds_to_timedelta(data['MAX_engine_scan_time'])
        ]
        output_data.append(row)

    # Create csv, if required
    if csv_config['enabled']:
        write_to_csv(['Week','Scans','No Scans','Full Scans','Incremental Scans','Sum LOC','Max LOC','Sum Failed LOC','Max Failed LOC',
            'AVG Total Scan Time','Max Total Scan Time','Avg Source Pulling Time','Max Source Pulling Time','Avg Queue Time','Max Queue Time',
            'Avg Engine Time','Max Engine Time'], output_data, os.path.join(csv_config['output_dir'], f'13-scans_by_week.csv'))

    # Output to Excel, if required
    if excel_config['enabled']:
        write_to_excel(excel_config['workbook']['Data'], output_data, 'BO', 4, _SCANS_BY_PERIOD_FORMATS)


def _entity_row(name, stats, extra_leading_cols):
    row = list(extra_leading_cols) + [
        stats['COUNT_scans'], stats['COUNT_full_scans'], stats['COUNT_incremental_scans'], stats['PCT_incremental'],
        stats['SUM_loc'], stats['AVG_loc'], stats['MAX_loc'], stats['SUM_failed_loc'], stats['MAX_failed_loc'],
        stats['AVG_file_count'], stats['MAX_file_count'],
        stats['first_scan_date'], stats['last_scan_date'], stats['active_day_count'], stats['AVG_scans_per_week'],
    ]
    for sev in ENTITY_SEVERITIES:
        row += [stats[f'AVG_{sev}'], stats[f'MAX_{sev}'], stats[f'MIN_{sev}']]
    return row


# Shared column shape of output_scans_by_date/output_scans_by_week: Date/Week, then
# Scans/NoScans/FullScans/IncrementalScans/LOC sums+maxes (int), then 4 pairs of
# avg/max duration values.
_SCANS_BY_PERIOD_FORMATS = (
    [cx_theme.FMT_DATE] + [cx_theme.FMT_INT] * 8 + [cx_theme.FMT_DURATION] * 8
)

_ENTITY_HEADER_TAIL = ['Scans', 'Full Scans', 'Incremental Scans', '% Incremental',
                       'Total LOC', 'Avg LOC/Scan', 'Max LOC', 'Total Failed LOC', 'Max Failed LOC',
                       'Avg File Count', 'Max File Count', 'First Scan', 'Last Scan', 'Active Days',
                       'Avg Scans/Week']
for _sev in ['Critical', 'High', 'Medium', 'Low', 'Info']:
    _ENTITY_HEADER_TAIL += [f'{_sev} Avg', f'{_sev} Max', f'{_sev} Min']

# Number formats for the entity (project/team) detail tables' shared tail columns
# (everything after the identity columns, which vary between Projects and Teams).
_ENTITY_FORMAT_TAIL = (
    [cx_theme.FMT_INT, cx_theme.FMT_INT, cx_theme.FMT_INT, cx_theme.FMT_PCT,
     cx_theme.FMT_INT, cx_theme.FMT_INT, cx_theme.FMT_INT, cx_theme.FMT_INT, cx_theme.FMT_INT,
     cx_theme.FMT_INT, cx_theme.FMT_INT, cx_theme.FMT_DATE, cx_theme.FMT_DATE, cx_theme.FMT_INT,
     cx_theme.FMT_DECIMAL2]
    + [cx_theme.FMT_DECIMAL2, cx_theme.FMT_INT, cx_theme.FMT_INT] * len(ENTITY_SEVERITIES)
)


def output_project_stats(project_stats, csv_config, excel_config):
    output_data = []
    for name, stats in sorted(project_stats.items(), key=lambda kv: kv[1]['COUNT_scans'], reverse=True):
        output_data.append(_entity_row(name, stats, [name, stats['team_name']]))

    if csv_config['enabled']:
        header = ['Logical Project', 'Team'] + _ENTITY_HEADER_TAIL
        write_to_csv(header, output_data, os.path.join(csv_config['output_dir'], '14-project_stats.csv'))

    if excel_config['enabled']:
        col_info = excel_config['proj_cols']
        ws = excel_config['workbook']['Projects']
        formats = [None, None] + _ENTITY_FORMAT_TAIL
        write_to_excel(ws, output_data, get_column_letter(col_info['identity_col_start']), col_info['detail_data_start'], formats)


def output_team_stats(team_stats, csv_config, excel_config):
    output_data = []
    for name, stats in sorted(team_stats.items(), key=lambda kv: kv[1]['COUNT_scans'], reverse=True):
        unique_projects = len(stats['unique_projects'])
        avg_scans_per_project = stats['COUNT_scans'] / unique_projects if unique_projects else 0
        output_data.append(_entity_row(name, stats, [name, unique_projects, avg_scans_per_project]))

    if csv_config['enabled']:
        header = ['Team', 'Unique Projects', 'Avg Scans/Project'] + _ENTITY_HEADER_TAIL
        write_to_csv(header, output_data, os.path.join(csv_config['output_dir'], '15-team_stats.csv'))

    if excel_config['enabled']:
        col_info = excel_config['team_cols']
        ws = excel_config['workbook']['Teams']
        formats = [None, cx_theme.FMT_INT, cx_theme.FMT_DECIMAL2] + _ENTITY_FORMAT_TAIL
        write_to_excel(ws, output_data, get_column_letter(col_info['identity_col_start']), col_info['detail_data_start'], formats)


def output_top_projects(project_stats, csv_config, excel_config):
    by_volume = sorted(project_stats.items(), key=lambda kv: kv[1]['COUNT_scans'], reverse=True)[:TOP_N_ENTITIES]
    by_size = sorted(project_stats.items(), key=lambda kv: kv[1]['SUM_loc'], reverse=True)[:TOP_N_ENTITIES]
    vol_data = [[name, s['team_name'], s['COUNT_scans']] for name, s in by_volume]
    size_data = [[name, s['team_name'], s['SUM_loc']] for name, s in by_size]

    if csv_config['enabled']:
        write_to_csv(['Logical Project', 'Team', 'Scans'], vol_data,
                     os.path.join(csv_config['output_dir'], '16-top_projects_by_volume.csv'))
        write_to_csv(['Logical Project', 'Team', 'Total LOC'], size_data,
                     os.path.join(csv_config['output_dir'], '17-top_projects_by_size.csv'))

    if excel_config['enabled']:
        col_info = excel_config['proj_cols']
        ws = excel_config['workbook']['Projects']
        write_to_excel(ws, vol_data, get_column_letter(col_info['vol_col_start']), col_info['vol_data_start'],
                        [None, None, cx_theme.FMT_INT])
        write_to_excel(ws, size_data, get_column_letter(col_info['size_col_start']), col_info['size_data_start'],
                        [None, None, cx_theme.FMT_INT])


def output_top_teams(team_stats, csv_config, excel_config):
    by_volume = sorted(team_stats.items(), key=lambda kv: kv[1]['COUNT_scans'], reverse=True)[:TOP_N_ENTITIES]
    by_size = sorted(team_stats.items(), key=lambda kv: kv[1]['SUM_loc'], reverse=True)[:TOP_N_ENTITIES]
    vol_data = [[name, s['COUNT_scans']] for name, s in by_volume]
    size_data = [[name, s['SUM_loc']] for name, s in by_size]

    if csv_config['enabled']:
        write_to_csv(['Team', 'Scans'], vol_data, os.path.join(csv_config['output_dir'], '18-top_teams_by_volume.csv'))
        write_to_csv(['Team', 'Total LOC'], size_data, os.path.join(csv_config['output_dir'], '19-top_teams_by_size.csv'))

    if excel_config['enabled']:
        col_info = excel_config['team_cols']
        ws = excel_config['workbook']['Teams']
        write_to_excel(ws, vol_data, get_column_letter(col_info['vol_col_start']), col_info['vol_data_start'],
                        [None, cx_theme.FMT_INT])
        write_to_excel(ws, size_data, get_column_letter(col_info['size_col_start']), col_info['size_data_start'],
                        [None, cx_theme.FMT_INT])


### Compute mean/median/stdev/min/p25/p75/max for one list of values
def _percentile(sorted_values, pct):
    if not sorted_values:
        return 0
    k = (len(sorted_values) - 1) * pct
    f, c = math.floor(k), math.ceil(k)
    if f == c:
        return sorted_values[int(k)]
    return sorted_values[f] * (c - k) + sorted_values[c] * (k - f)


def _summary_stats(values):
    values = sorted(v for v in values if v is not None)
    if not values:
        return {'mean': 0, 'median': 0, 'stdev': 0, 'min': 0, 'p25': 0, 'p75': 0, 'max': 0}
    return {
        'mean': statistics.mean(values),
        'median': statistics.median(values),
        'stdev': statistics.stdev(values) if len(values) > 1 else 0,
        'min': values[0],
        'p25': _percentile(values, 0.25),
        'p75': _percentile(values, 0.75),
        'max': values[-1],
    }


### Bin a list of values into a histogram. Uses log-scaled bins when the data spans
# more than two orders of magnitude (typical for scan volume/LOC, where a handful of
# huge outliers would otherwise collapse everything else into a single linear bin).
def _histogram_bins(values, num_bins=10):
    values = [v for v in values if v is not None]
    if not values:
        return [("(no data)", 0)] + [("", 0)] * (num_bins - 1)
    lo, hi = min(values), max(values)
    if lo == hi:
        return [(f"{lo:,.0f}", len(values))] + [("", 0)] * (num_bins - 1)

    use_log = lo > 0 and hi / lo > 100
    if use_log:
        log_lo, log_hi = math.log10(lo), math.log10(hi)
        edges = [10 ** (log_lo + (log_hi - log_lo) * i / num_bins) for i in range(num_bins + 1)]
    else:
        edges = [lo + (hi - lo) * i / num_bins for i in range(num_bins + 1)]

    counts = [0] * num_bins
    for v in values:
        idx = max(0, min(bisect.bisect_right(edges, v) - 1, num_bins - 1))
        counts[idx] += 1

    labels = [f"{edges[i]:,.0f}-{edges[i + 1]:,.0f}" for i in range(num_bins)]
    return list(zip(labels, counts))


def _entity_distribution_values(stats_dict):
    return {
        'Scans': [s['COUNT_scans'] for s in stats_dict.values()],
        'Total LOC': [s['SUM_loc'] for s in stats_dict.values()],
        'Avg LOC/Scan': [s['AVG_loc'] for s in stats_dict.values()],
        '% Incremental': [s['PCT_incremental'] for s in stats_dict.values()],
        'Active Days': [s['active_day_count'] for s in stats_dict.values()],
        'Avg Scans/Week': [s['AVG_scans_per_week'] for s in stats_dict.values()],
        'Critical Avg': [s['AVG_critical'] for s in stats_dict.values()],
        'High Avg': [s['AVG_high'] for s in stats_dict.values()],
        'Medium Avg': [s['AVG_medium'] for s in stats_dict.values()],
        'Low Avg': [s['AVG_low'] for s in stats_dict.values()],
        'Info Avg': [s['AVG_info'] for s in stats_dict.values()],
    }


def _output_distribution(stats_dict, csv_config, excel_config, sheet_name, col_info_key, csv_filenames):
    values_by_metric = _entity_distribution_values(stats_dict)

    stats_rows = []
    for metric in workbook_builder.DISTRIBUTION_METRICS:
        s = _summary_stats(values_by_metric[metric])
        stats_rows.append([metric, s['mean'], s['median'], s['stdev'], s['min'], s['p25'], s['p75'], s['max']])

    scans_bins = [list(b) for b in _histogram_bins(values_by_metric['Scans'])]
    loc_bins = [list(b) for b in _histogram_bins(values_by_metric['Total LOC'])]

    if csv_config['enabled']:
        write_to_csv(workbook_builder.DISTRIBUTION_STAT_COLS, stats_rows,
                     os.path.join(csv_config['output_dir'], csv_filenames[0]))
        write_to_csv(['Range', 'Count'], scans_bins, os.path.join(csv_config['output_dir'], csv_filenames[1]))
        write_to_csv(['Range', 'Count'], loc_bins, os.path.join(csv_config['output_dir'], csv_filenames[2]))

    if excel_config['enabled']:
        col_info = excel_config[col_info_key]
        ws = excel_config['workbook'][sheet_name]
        D2 = cx_theme.FMT_DECIMAL2
        write_to_excel(ws, stats_rows, get_column_letter(col_info['dist_col_start']),
                        col_info['dist_stats_data_start'], [None, D2, D2, D2, D2, D2, D2, D2])
        write_to_excel(ws, scans_bins, get_column_letter(col_info['dist_col_start']),
                        col_info['dist_scans_hist_data_start'], [None, cx_theme.FMT_INT])
        write_to_excel(ws, loc_bins, get_column_letter(col_info['dist_col_start']),
                        col_info['dist_loc_hist_data_start'], [None, cx_theme.FMT_INT])


def output_project_distribution(project_stats, csv_config, excel_config):
    _output_distribution(project_stats, csv_config, excel_config, 'Projects', 'proj_cols',
                          ['20-project_distribution_stats.csv', '21-project_scans_histogram.csv',
                           '22-project_loc_histogram.csv'])


def output_team_distribution(team_stats, csv_config, excel_config):
    _output_distribution(team_stats, csv_config, excel_config, 'Teams', 'team_cols',
                          ['23-team_distribution_stats.csv', '24-team_scans_histogram.csv',
                           '25-team_loc_histogram.csv'])


### Main
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process scans and output CSV files if requested")
    parser.add_argument("input_file", type=str, help="JSON file containing scan data")
    parser.add_argument("--customer", type=str, default="", help="Optional name of the customer")
    parser.add_argument("--cc-snapshot", type=int, default=CC_SNAPSHOT_SECONDS, help="Interval in seconds for capturing concurrency snapshots (default: %(default)s)")
    parser.add_argument("--csv", action="store_true", help="Generate CSV output files")
    parser.add_argument("--full-data", action="store_true", help="Generate CSV output of complete scan data")
    parser.add_argument("--excel", action="store_true", help="Generate an Excel (.xlsx) report -- built from scratch, no template file needed")

    args = parser.parse_args()

    # Make sure that some sort of output is defined
    if not (args.csv or args.full_data or args.excel):
        parser.error("At least one of the output options --csv, --full-data, or --excel must be specified.")
        exit(1)

    input_file = args.input_file
    output_name = args.customer.replace(" ", "_") if args.customer else os.path.splitext(os.path.basename(input_file))[0]

    if args.cc_snapshot:
        CC_SNAPSHOT_SECONDS = args.cc_snapshot

    # Define the output directory using the optional name if provided
    start_time = datetime.now().strftime('%Y%m%d-%H%M%S')
    output_dir = os.path.join(os.getcwd(), f"ehc_output_{output_name}_{start_time}")

    # Define the target Excel workbook filename
    if args.excel:
        excel_filename = f"EHC-{output_name}.xlsx"
        excel_target_full_path = os.path.join(output_dir, excel_filename)
    else:
        excel_filename = None
        excel_target_full_path = None

    # Initialize structures to hold output configs
    full_csv = {
        'enabled': True if args.full_data else False,
        'output_dir': output_dir,
        'field_names': []
    }
    csv_config = {
        'enabled': True if args.csv else False,
        'output_dir': output_dir
    }
    excel_config = {
        'enabled': True if args.excel else False,
        'excel_target': excel_target_full_path,
        'workbook': None,
        'proj_cols': None,
        'team_cols': None,
    }

    # If we are creating any files, create the output directory
    if args.full_data or args.csv or args.excel:
        try:
            # Attempt to create the directory
            os.makedirs(output_dir, exist_ok=True)
        except PermissionError as e:
            print(f"Permission Error: {e}")
            exit(1)
        except Exception as e:
            print(f"Error creating directory: {e}")
            exit(1)

    # If we're exporting to Excel, build the workbook from scratch (no template file involved)
    if excel_config['enabled']:
        if not xl_available:
            parser.error("--excel requires the openpyxl and Pillow libraries: 'pip install openpyxl Pillow'")
        workbook, proj_cols, team_cols = workbook_builder.create_workbook()
        excel_config['workbook'] = workbook
        excel_config['proj_cols'] = proj_cols
        excel_config['team_cols'] = team_cols

    full_csv['field_names'], scans = ingest_file(input_file)

    processed_data = process_scans(scans, full_csv)

    output_analysis(processed_data, csv_config, excel_config)

    # If we exported to Excel, save the workbook
    if excel_config['enabled']:
        workbook_builder.save_workbook(excel_config['workbook'], excel_target_full_path)

    end_time = datetime.now().strftime('%Y%m%d-%H%M%S')
    elapsed_time = format_seconds_to_hms(calculate_time_difference(start_time, end_time))

    print (f"FastEHC analyzed {processed_data['aggregate_metrics']['COUNT_scans']:,} scans in {elapsed_time} and wrote output files to {output_dir}")
