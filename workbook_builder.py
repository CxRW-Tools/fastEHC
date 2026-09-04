"""Builds the fastEHC output workbook from scratch (no external .xlsx template).

Each sheet is constructed with its static labels, formulas, charts, and Checkmarx
branding already in place; the `output_*` functions in fastEHC.py then fill in
values via `write_to_excel()`, which styles each cell as it writes it.
"""
import os
import re
import zipfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

import cx_theme as theme

_DRAWINGML_NS = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
_ANCHOR_BLOCK_RE = re.compile(r"<(oneCellAnchor|twoCellAnchor)>.*?</\1>", re.DOTALL)
_EXT_RE = re.compile(r'<ext cx="(\d+)" cy="(\d+)"/>')


def _patch_anchor_block(match):
    """openpyxl anchors every chart's graphicFrame with an empty <xfrm/> (no off/ext
    children) when the chart was added via a plain cell-string anchor. That's schema-
    invalid -- Excel silently discards the entire drawing part (all charts *and* any
    image sharing that sheet) rather than just the offending chart. Fill it in using
    the same extent the (valid) outer anchor already carries."""
    block = match.group(0)
    if "<graphicFrame>" not in block or "<xfrm/>" not in block:
        return block
    ext_match = _EXT_RE.search(block)
    if not ext_match:
        return block
    cx, cy = ext_match.groups()
    replacement = (f'<xfrm><a:off {_DRAWINGML_NS} x="0" y="0"/>'
                   f'<a:ext {_DRAWINGML_NS} cx="{cx}" cy="{cy}"/></xfrm>')
    return block.replace("<xfrm/>", replacement, 1)


def save_workbook(wb, path):
    """Save the workbook, then patch the empty-<xfrm/> defect openpyxl 3.1.5 leaves
    on every chart's graphicFrame (see _patch_anchor_block)."""
    wb.save(path)

    tmp_path = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if re.match(r"xl/drawings/drawing\d+\.xml$", item.filename):
                text = data.decode("utf-8")
                text = _ANCHOR_BLOCK_RE.sub(_patch_anchor_block, text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp_path, path)

# ---- Column groups shared by the Projects/Teams detail tables ----
SEVERITIES = ["Critical", "High", "Medium", "Low", "Info"]
SEV_SUBCOLS = ["Avg", "Max", "Min"]
VOLUME_COLS = ["Scans", "Full Scans", "Incremental Scans", "% Incremental"]
SIZE_COLS = ["Total LOC", "Avg LOC/Scan", "Max LOC", "Total Failed LOC", "Max Failed LOC",
             "Avg File Count", "Max File Count"]
ACTIVITY_COLS = ["First Scan", "Last Scan", "Active Days", "Avg Scans/Week"]

# Metrics covered by the "Overall Distribution" stats table (mean/median/etc across
# all projects or all teams) -- matches the keys fastEHC.py's _distribution_stats()
# produces, in display order.
DISTRIBUTION_METRICS = ["Scans", "Total LOC", "Avg LOC/Scan", "% Incremental", "Active Days",
                         "Avg Scans/Week", "Critical Avg", "High Avg", "Medium Avg", "Low Avg", "Info Avg"]
DISTRIBUTION_STAT_COLS = ["Metric", "Mean", "Median", "Std Dev", "Min", "P25", "P75", "Max"]
HISTOGRAM_BIN_COUNT = 10

# openpyxl defaults every Bar/LineChart's axis IDs to the same values (10/100),
# which silently corrupts multiple such charts sharing a workbook on save/reload
# -- the axIds collide and one chart's series data gets dropped entirely. Every
# bar/line chart (or combo pair) needs a distinct axId.
_id_counter = [10]


def _next_id():
    _id_counter[0] += 1
    return _id_counter[0]


def _assign_primary_axes(chart):
    chart.x_axis.axId = _next_id()
    chart.y_axis.axId = _next_id()


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb)
    proj_cols = _build_projects_or_teams_sheet(wb, "Projects", identity_cols=["Logical Project", "Team"])
    team_cols = _build_projects_or_teams_sheet(
        wb, "Teams", identity_cols=["Team", "Unique Projects", "Avg Scans/Project"])
    _build_scan_time_analysis_sheet(wb)
    _build_data_sheet(wb)
    _build_charts_sheet(wb, proj_cols, team_cols)

    # Data must exist before Charts is built (charts reference it), but the tab
    # order we actually want to present is: Summary, Projects, Teams, Scan Time
    # Analysis, Charts, Data (raw backing sheet last).
    desired_order = ["Summary", "Projects", "Teams", "Scan Time Analysis", "Charts", "Data"]
    wb._sheets = [wb[name] for name in desired_order]

    wb.active = 0
    return wb, proj_cols, team_cols


# ======================= DATA SHEET =======================

DATA_SECTIONS = [
    ("B", "01-summary_of_scans.csv", ["Description", "Value", "%"]),
    ("F", "02-scan_metrics.csv", ["Description", "Average", "Max"]),
    ("J", "03-scan_duration.csv", ["Description", "Average", "Max"]),
    ("N", "04-scan_results_severity.csv", ["Description", "Average", "Max"]),
    ("R", "05-languages.csv", ["Language", "%", "Scans"]),
    ("V", "06-scan_submissison_summary.csv", ["Description", "Value"]),
    ("Y", "07-day_of_week_scan_average.csv", ["Day of Week", "Scans", "%"]),
    ("AC", "08-origins.csv", ["Origin", "Scans", "%"]),
    ("AG", "09-presets.csv", ["Preset", "Scans", "%"]),
    ("AK", "10-scan_time_analysis.csv", ["LOC Range", "Scans", "% Scans", "Avg Total Time",
                                          "Avg Source Pulling Time", "Avg Queue", "Avg Engine"]),
    ("AS", "11-concurrency_analysis.csv", ["Date", "Max Actual", "Max Optimal"]),
    ("AW", "12-scans_by_date.csv", ["Date", "Scans", "No Scans", "Full Scans", "Incremental Scans",
                                     "Sum LOC", "Max LOC", "Sum Failed LOC", "Max Failed LOC",
                                     "AVG Total Scan Time", "Max Total Scan Time",
                                     "Avg Source Pulling Time", "Max Source Pulling Time",
                                     "Avg Queue Time", "Max Queue Time", "Avg Engine Time", "Max Engine Time"]),
    ("BO", "13-scans_by_week.csv", ["Week", "Scans", "No Scans", "Full Scans", "Incremental Scans",
                                     "Sum LOC", "Max LOC", "Sum Failed LOC", "Max Failed LOC",
                                     "AVG Total Scan Time", "Max Total Scan Time",
                                     "Avg Source Pulling Time", "Max Source Pulling Time",
                                     "Avg Queue Time", "Max Queue Time", "Avg Engine Time", "Max Engine Time"]),
]


def _build_data_sheet(wb):
    ws = wb.create_sheet("Data")
    theme.add_corner_logo(ws)
    ws.sheet_properties.tabColor = theme.QUANTUM_VIOLET

    for start_col, csv_name, headers in DATA_SECTIONS:
        col0 = _col_idx(start_col)
        note_cell = ws.cell(row=2, column=col0, value=csv_name)
        theme.style_note_cell(note_cell)
        for i, h in enumerate(headers):
            theme.style_header_cell(ws.cell(row=3, column=col0 + i, value=h))

    # Severity labels are always this fixed list -- pre-fill + color now so the
    # rows read correctly even before output_scan_results_and_severity() runs,
    # and stay colored once it does (same literal text is written either way).
    n_col = _col_idx("N")
    for offset, (label, sev) in enumerate(zip(
            ["Total", "Critical", "High", "Medium", "Low", "Informational"],
            [None, "critical", "high", "medium", "low", "informational"])):
        cell = ws.cell(row=4 + offset, column=n_col, value=label)
        if sev:
            font, fill = theme.severity_fill_font(sev)
            cell.font, cell.fill = font, fill
        else:
            theme.style_body_cell(cell)

    ws.freeze_panes = "B4"
    return ws


def _col_idx(letters):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letters)


# ======================= SUMMARY SHEET =======================

def _build_summary_sheet(wb):
    ws = wb.create_sheet("Summary")
    theme.add_corner_logo(ws)
    ws.sheet_properties.tabColor = theme.QUANTUM_VIOLET

    def group(cell_range, text):
        cell = ws[cell_range.split(":")[0]]
        cell.value = text
        theme.style_group_cell(cell)

    def header(coord, text):
        theme.style_header_cell(ws[coord])
        ws[coord] = text

    def formula(coord, f, number_format=None):
        cell = ws[coord]
        theme.style_body_cell(cell)
        cell.value = f
        if number_format:
            cell.number_format = number_format

    FI, FP, FDT, FDU, FD2 = (theme.FMT_INT, theme.FMT_PCT, theme.FMT_DATE, theme.FMT_DURATION,
                              theme.FMT_DECIMAL2)

    # ---- top-level section banners ----
    group("B2", "Scan Metrics")
    group("F2", "Language & Size")
    group("J2", "Scanning Behavior")
    group("N2", "Results")

    # ---- Scans Summary (B4:D21) ----
    for coord, text in [("B4", "Scans Summary"), ("C4", "Values"), ("D4", "%")]:
        header(coord, text)
    labels = ["Start Date", "End Date", "Days", "Weeks", "Scans Submitted", "Scans Completed",
              "Scans Failed", "Full Scans Submitted", "Incremental Scans Submitted", "No-Change Scans",
              "Scans with Critical Results", "Scans with High Results", "Scans with Medium Results",
              "Scans with Low Results", "Scans with Informational Results", "Scans with Zero Results",
              "Unique Projects Scanned"]
    # Rows 5-6 (Start/End Date) get a date format; every other row here is a plain count.
    for i, label in enumerate(labels):
        row = 5 + i
        theme.style_body_cell(ws[f"B{row}"])
        ws[f"B{row}"] = label
        formula(f"C{row}", f"=Data!C{row - 1}", FDT if row in (5, 6) else FI)
    formula("D10", "=C10/C9", FP)
    formula("D11", "=C11/C9", FP)
    for row in range(12, 20):
        formula(f"D{row}", f"=Data!D{row - 1}", FP)

    # ---- Language & Size (F4:H26) ----
    for coord, text in [("F4", "Language"), ("G4", "# of Scans "), ("H4", "% Scans")]:
        header(coord, text)
    languages = ["Apex", "ASP", "Cobol", "CPP", "CSharp", "Go", "Groovy", "Java", "JavaScript",
                 "Kotlin", "Objc", "Perl", "PHP", "PLSQL", "Python", "Ruby", "Scala", "VB6", "VbNet",
                 "VbScript", "Unknown", "Typescript"]
    for i, lang in enumerate(languages):
        row = 5 + i
        theme.style_body_cell(ws[f"F{row}"])
        ws[f"F{row}"] = lang
        formula(f"G{row}", f'=_xlfn.IFNA(VLOOKUP($F{row},Data!$R$4:$T$28,3,FALSE),"")', FI)
        formula(f"H{row}", f'=_xlfn.IFNA(VLOOKUP($F{row},Data!$R$4:$T$28,2,FALSE),"")', FP)

    # ---- Scanning Behavior: Scan Submission Summary (J4:K10) ----
    for coord, text in [("J4", "Scan Submission Summary"), ("K4", "Values")]:
        header(coord, text)
    submission_labels = ["Average Scans Submitted per Week", "Average Scans Submitted per Day",
                          "Average Scans Submitted per Week Day", "Average Scans Submitted per Weekend Day",
                          "Max Daily Scans Submitted", "Date of Max Scans"]
    for i, label in enumerate(submission_labels):
        row = 5 + i
        theme.style_body_cell(ws[f"J{row}"])
        ws[f"J{row}"] = label
        fmt = FDT if row == 10 else (FI if row == 9 else FD2)
        formula(f"K{row}", f"=Data!W{row - 1}", fmt)

    # ---- Day of Week Scan Average (J12:L19) ----
    header("J12", "Day of Week Scan Average")
    header("K12", "# of Scans ")
    header("L12", "% Scans")
    for i, day in enumerate(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]):
        row = 13 + i
        theme.style_body_cell(ws[f"J{row}"])
        ws[f"J{row}"] = day
        formula(f"K{row}", f"=Data!Z{row - 9}", FI)
        formula(f"L{row}", f"=Data!AA{row - 9}", FP)

    # ---- Origin (J21:L36) ----
    header("J21", "Origin")
    header("K21", "# of Scans ")
    header("L21", "% Scans")
    for row in range(22, 37):
        data_row = row - 18
        formula(f"J{row}", f'=IF(ISBLANK(Data!AC{data_row}),"",Data!AC{data_row})')
        formula(f"K{row}", f'=IF(ISBLANK(Data!AD{data_row}),"",Data!AD{data_row})', FI)
        formula(f"L{row}", f'=IF(ISBLANK(Data!AE{data_row}),"",Data!AE{data_row})', FP)

    # ---- Results: Scan Results / Severity (N4:P10) ----
    for coord, text in [("N4", "Scan Results / Severity"), ("O4", "Average"), ("P4", "Max")]:
        header(coord, text)
    result_labels = ["Total", "Critical", "High", "Medium", "Low", "Informational"]
    for i, label in enumerate(result_labels):
        row = 5 + i
        theme.style_body_cell(ws[f"N{row}"])
        ws[f"N{row}"] = label
        formula(f"O{row}", f"=Data!O{row - 1}", FI)
        formula(f"P{row}", f"=Data!P{row - 1}", FI)

    # ---- Preset Selection (N12:P27) ----
    header("N12", "Preset Selection")
    header("O12", "# of Scans ")
    header("P12", "% Scans")
    for row in range(13, 28):
        data_row = row - 9
        formula(f"N{row}", f'=IF(ISBLANK(Data!AG{data_row}),"",Data!AG{data_row})')
        formula(f"O{row}", f'=IF(ISBLANK(Data!AH{data_row}),"",Data!AH{data_row})', FI)
        formula(f"P{row}", f'=IF(ISBLANK(Data!AI{data_row}),"",Data!AI{data_row})', FP)

    # ---- Scan Metrics (B23:D26) ----
    for coord, text in [("B23", "Scan Metrics"), ("C23", "Average"), ("D23", "Max")]:
        header(coord, text)
    for i, label in enumerate(["LOC per Scan", "Failed LOC per Scan", "Daily LOC"]):
        row = 24 + i
        theme.style_body_cell(ws[f"B{row}"])
        ws[f"B{row}"] = label
        formula(f"C{row}", f"=Data!G{row - 20}", FI)
        formula(f"D{row}", f"=Data!H{row - 20}", FI)

    # ---- Scan Duration (B28:D32) ----
    for coord, text in [("B28", "Scan Duration"), ("C28", "Average"), ("D28", "Max")]:
        header(coord, text)
    for i, label in enumerate(["Total Scan Duration", "Source Pulling Duration", "Queued Duration",
                                "Engine Scan Duration"]):
        row = 29 + i
        theme.style_body_cell(ws[f"B{row}"])
        ws[f"B{row}"] = label
        formula(f"C{row}", f"=Data!K{row - 25}", FDU)
        formula(f"D{row}", f"=Data!L{row - 25}", FDU)

    # ---- LOC Range (F28:H40) ----
    for coord, text in [("F28", "LOC Range"), ("G28", "# of Scans "), ("H28", "% Scans")]:
        header(coord, text)
    loc_ranges = ["0 to 20k", "20k-50k", "50k-100k", "100k-250k", "250k-500k", "500k-1M",
                  "1M-2M", "2M-3M", "3M-5M", "5M-7M", "7M-10M", "10M+"]
    for i, rng in enumerate(loc_ranges):
        row = 29 + i
        theme.style_body_cell(ws[f"F{row}"])
        ws[f"F{row}"] = rng
        formula(f"G{row}", f"=Data!AL{row - 25}", FI)
        formula(f"H{row}", f"=Data!AM{row - 25}", FP)

    for col in "BCDFGHJKLNOP":
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["J"].width = 40
    ws.column_dimensions["N"].width = 26
    ws.freeze_panes = "B5"
    return ws


# ======================= SCAN TIME ANALYSIS SHEET =======================

def _build_scan_time_analysis_sheet(wb):
    ws = wb.create_sheet("Scan Time Analysis")
    theme.add_corner_logo(ws)
    ws.sheet_properties.tabColor = theme.QUANTUM_VIOLET

    theme.style_group_cell(ws["B2"])
    ws["B2"] = "Scan Time Analysis"

    headers = ["LOC Range", "Scans", "% Scans", "Avg Total Time", "Avg Source Pulling Time",
               "Avg Queue Time", "Avg Engine Scan Time"]
    for i, h in enumerate(headers):
        theme.style_header_cell(ws.cell(row=4, column=2 + i, value=h))

    data_cols = ["AK", "AL", "AM", "AN", "AO", "AP", "AQ"]
    col_formats = [None, theme.FMT_INT, theme.FMT_PCT, theme.FMT_DURATION, theme.FMT_DURATION,
                   theme.FMT_DURATION, theme.FMT_DURATION]
    for i in range(12):  # 12 LOC-range rows
        row = 5 + i
        for j, dcol in enumerate(data_cols):
            col_letter = get_column_letter(2 + j)
            cell = ws[f"{col_letter}{row}"]
            theme.style_body_cell(cell)
            cell.value = f"=Data!{dcol}{row - 1}"
            if col_formats[j]:
                cell.number_format = col_formats[j]
    theme.style_body_cell(ws["C17"])
    ws["C17"] = "=SUM(C5:C16)"
    ws["C17"].number_format = theme.FMT_INT

    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Scans by LOC Range"
    cats = Reference(ws, min_col=2, min_row=5, max_row=16)
    vals = Reference(ws, min_col=3, min_row=4, max_row=16)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    theme.color_chart_series(chart)
    _assign_primary_axes(chart)
    chart.height, chart.width = 9, 18
    ws.add_chart(chart, "J4")
    return ws


# ======================= CHARTS SHEET =======================

def _combo_bar_line(ws, bar_ref_col, bar_cat_col, bar_name_row, first_data_row, last_data_row,
                     line_cols, title):
    bar = BarChart()
    bar.type = "col"
    bar.title = title
    cats = Reference(ws.parent["Data"], min_col=_col_idx(bar_cat_col), min_row=first_data_row, max_row=last_data_row)
    vals = Reference(ws.parent["Data"], min_col=_col_idx(bar_ref_col), min_row=bar_name_row, max_row=last_data_row)
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    theme.color_chart_series(bar, start_index=0)
    _assign_primary_axes(bar)

    line = LineChart()
    for i, col in enumerate(line_cols):
        vals = Reference(ws.parent["Data"], min_col=_col_idx(col), min_row=bar_name_row, max_row=last_data_row)
        line.add_data(vals, titles_from_data=True)
    theme.color_chart_series(line, start_index=1)
    # Combo chart: share the bar's category axis, give the line its own secondary
    # value axis (time durations and scan counts are very different scales).
    line.x_axis.axId = bar.x_axis.axId
    line.y_axis.axId = _next_id()
    bar.y_axis.crosses = "max"
    bar += line
    bar.height, bar.width = 9, 30
    return bar


def _build_charts_sheet(wb, proj_cols, team_cols):
    ws = wb.create_sheet("Charts")
    theme.add_corner_logo(ws)
    ws.sheet_properties.tabColor = theme.QUANTUM_VIOLET

    def banner(rng, text):
        cell = ws[rng.split(":")[0]]
        cell.value = text
        theme.style_group_cell(cell, horizontal="center")
        ws.merge_cells(rng)

    banner("B2:R2", "Charts")

    # --- Daily Scan Summary (bar: Scans, line: Avg Source Pulling / Queue Time) ---
    banner("B4:R4", "Daily Scan Summary")
    chart = _combo_bar_line(ws, "AX", "AW", 3, 4, 800, ["BH", "BJ"], "Daily Scan Summary")
    ws.add_chart(chart, "B5")

    # --- Weekly Scan Summary ---
    banner("B28:R28", "Weekly Scan Summary")
    chart = _combo_bar_line(ws, "BP", "BO", 3, 4, 115, ["BZ", "CB"], "Weekly Scan Summary")
    ws.add_chart(chart, "B29")

    # --- Concurrency Analysis ---
    banner("B52:R52", "Concurrency Analysis")
    line = LineChart()
    line.title = "Concurrency Analysis"
    data_ws = wb["Data"]
    cats = Reference(data_ws, min_col=_col_idx("AS"), min_row=4, max_row=92)
    for col in ("AT", "AU"):
        vals = Reference(data_ws, min_col=_col_idx(col), min_row=3, max_row=92)
        line.add_data(vals, titles_from_data=True)
    line.set_categories(cats)
    theme.color_chart_series(line)
    _assign_primary_axes(line)
    line.height, line.width = 9, 30
    ws.add_chart(line, "B53")

    # --- Language Analysis ---
    banner("B76:R76", "Language Analysis")
    bar = BarChart()
    bar.title = "Scan Languages"
    cats = Reference(data_ws, min_col=_col_idx("R"), min_row=4, max_row=21)
    vals = Reference(data_ws, min_col=_col_idx("T"), min_row=3, max_row=21)
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    theme.color_chart_series(bar)
    _assign_primary_axes(bar)
    bar.height, bar.width = 9, 30
    ws.add_chart(bar, "B77")

    # --- Scan Size Analysis (pie: scans by LOC range) ---
    banner("B100:I100", "Scan Size Analysis")
    pie = PieChart()
    pie.title = "Scans by LOC Range"
    cats = Reference(data_ws, min_col=_col_idx("AK"), min_row=4, max_row=15)
    vals = Reference(data_ws, min_col=_col_idx("AL"), min_row=3, max_row=15)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    theme.color_chart_series(pie)
    pie.height, pie.width = 9, 14
    ws.add_chart(pie, "B101")

    # --- Origin Analysis ---
    banner("K100:R100", "Origin Analysis")
    pie = PieChart()
    pie.title = "Scan Origins"
    cats = Reference(data_ws, min_col=_col_idx("AC"), min_row=4, max_row=50)
    vals = Reference(data_ws, min_col=_col_idx("AD"), min_row=3, max_row=50)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    theme.color_chart_series(pie)
    pie.height, pie.width = 9, 14
    ws.add_chart(pie, "K101")

    # --- Preset Analysis ---
    banner("B124:I124", "Preset Analysis")
    pie = PieChart()
    pie.title = "Scan Presets"
    cats = Reference(data_ws, min_col=_col_idx("AG"), min_row=4, max_row=50)
    vals = Reference(data_ws, min_col=_col_idx("AH"), min_row=3, max_row=50)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    theme.color_chart_series(pie)
    pie.height, pie.width = 9, 14
    ws.add_chart(pie, "B125")

    # --- Results Analysis (severity pie, exact brand severity colors) ---
    banner("K124:R124", "Results Analysis")
    pie = PieChart()
    pie.title = "Average Results per Scan"
    cats = Reference(data_ws, min_col=_col_idx("N"), min_row=5, max_row=8)
    vals = Reference(data_ws, min_col=_col_idx("O"), min_row=5, max_row=8)
    pie.add_data(vals, titles_from_data=False)
    pie.set_categories(cats)
    theme.color_severity_pie(pie)
    pie.height, pie.width = 9, 14
    ws.add_chart(pie, "K125")

    # --- Project & Team Analysis ---
    banner("B148:R148", "Project & Team Analysis")
    proj_bar = BarChart()
    proj_bar.title = "Top Projects by Scan Volume"
    proj_ws = wb["Projects"]
    cats = Reference(proj_ws, min_col=proj_cols["identity"][0], min_row=proj_cols["vol_data_start"],
                      max_row=proj_cols["vol_data_end"])
    vals = Reference(proj_ws, min_col=proj_cols["vol_scans_col"], min_row=proj_cols["vol_header_row"],
                      max_row=proj_cols["vol_data_end"])
    proj_bar.add_data(vals, titles_from_data=True)
    proj_bar.set_categories(cats)
    theme.color_chart_series(proj_bar)
    _assign_primary_axes(proj_bar)
    proj_bar.height, proj_bar.width = 9, 22
    ws.add_chart(proj_bar, "B149")

    proj_bar2 = BarChart()
    proj_bar2.title = "Top Projects by Scan Size"
    cats = Reference(proj_ws, min_col=proj_cols["identity"][0], min_row=proj_cols["size_data_start"],
                      max_row=proj_cols["size_data_end"])
    vals = Reference(proj_ws, min_col=proj_cols["size_loc_col"], min_row=proj_cols["size_header_row"],
                      max_row=proj_cols["size_data_end"])
    proj_bar2.add_data(vals, titles_from_data=True)
    proj_bar2.set_categories(cats)
    theme.color_chart_series(proj_bar2)
    _assign_primary_axes(proj_bar2)
    proj_bar2.height, proj_bar2.width = 9, 22
    ws.add_chart(proj_bar2, "L149")

    team_ws = wb["Teams"]
    team_bar = BarChart()
    team_bar.title = "Teams by Scan Volume"
    cats = Reference(team_ws, min_col=team_cols["identity"][0], min_row=team_cols["vol_data_start"],
                      max_row=team_cols["vol_data_end"])
    vals = Reference(team_ws, min_col=team_cols["vol_scans_col"], min_row=team_cols["vol_header_row"],
                      max_row=team_cols["vol_data_end"])
    team_bar.add_data(vals, titles_from_data=True)
    team_bar.set_categories(cats)
    theme.color_chart_series(team_bar)
    _assign_primary_axes(team_bar)
    team_bar.height, team_bar.width = 9, 22
    ws.add_chart(team_bar, "B173")

    team_bar2 = BarChart()
    team_bar2.title = "Teams by Scan Size"
    cats = Reference(team_ws, min_col=team_cols["identity"][0], min_row=team_cols["size_data_start"],
                      max_row=team_cols["size_data_end"])
    vals = Reference(team_ws, min_col=team_cols["size_loc_col"], min_row=team_cols["size_header_row"],
                      max_row=team_cols["size_data_end"])
    team_bar2.add_data(vals, titles_from_data=True)
    team_bar2.set_categories(cats)
    theme.color_chart_series(team_bar2)
    _assign_primary_axes(team_bar2)
    team_bar2.height, team_bar2.width = 9, 22
    ws.add_chart(team_bar2, "L173")

    return ws


# ======================= PROJECTS / TEAMS SHEETS =======================

def _build_projects_or_teams_sheet(wb, sheet_name, identity_cols):
    ws = wb.create_sheet(sheet_name)
    theme.add_corner_logo(ws)
    ws.sheet_properties.tabColor = theme.QUANTUM_VIOLET

    entity = "Project" if sheet_name == "Projects" else "Team"
    title_cell = ws["B2"]
    title_cell.value = f"{sheet_name} — Scan Size, Volume & Severity by {'Logical Project' if entity == 'Project' else 'Team'}"
    title_cell.font = theme.TITLE_FONT

    TOP_N = 15
    col_info = {}

    # ---- Top N by Volume (left block) ----
    vol_header_row = 5
    theme.style_group_cell(ws.cell(row=vol_header_row - 1, column=2), )
    ws.cell(row=vol_header_row - 1, column=2, value=f"Top {TOP_N} {sheet_name} by Scan Volume")
    ws.merge_cells(start_row=vol_header_row - 1, start_column=2, end_row=vol_header_row - 1, end_column=4)
    vol_cols = [identity_cols[0]] + (["Team"] if entity == "Project" else []) + ["Scans"]
    for i, h in enumerate(vol_cols):
        theme.style_header_cell(ws.cell(row=vol_header_row, column=2 + i, value=h))
    vol_scans_col = 2 + len(vol_cols) - 1
    col_info.update(vol_header_row=vol_header_row, vol_data_start=vol_header_row + 1,
                     vol_data_end=vol_header_row + TOP_N, vol_scans_col=vol_scans_col, vol_col_start=2)

    # ---- Top N by Size (right block) ----
    size_start_col = 2 + len(vol_cols) + 2
    theme.style_group_cell(ws.cell(row=vol_header_row - 1, column=size_start_col))
    ws.cell(row=vol_header_row - 1, column=size_start_col, value=f"Top {TOP_N} {sheet_name} by Scan Size")
    ws.merge_cells(start_row=vol_header_row - 1, start_column=size_start_col,
                    end_row=vol_header_row - 1, end_column=size_start_col + len(vol_cols) - 1)
    size_cols = [identity_cols[0]] + (["Team"] if entity == "Project" else []) + ["Total LOC"]
    for i, h in enumerate(size_cols):
        theme.style_header_cell(ws.cell(row=vol_header_row, column=size_start_col + i, value=h))
    size_loc_col = size_start_col + len(size_cols) - 1
    col_info.update(size_header_row=vol_header_row, size_data_start=vol_header_row + 1,
                     size_data_end=vol_header_row + TOP_N, size_loc_col=size_loc_col,
                     size_col_start=size_start_col)

    # ---- Full detail table (grouped headers) ----
    detail_group_row = vol_header_row + TOP_N + 4
    detail_header_row = detail_group_row + 1
    col = 2
    group_spans = []

    def write_group(label, cols):
        nonlocal col
        start = col
        for c in cols:
            theme.style_header_cell(ws.cell(row=detail_header_row, column=col, value=c))
            col += 1
        group_spans.append((label, start, col - 1))

    write_group("Identity", identity_cols)
    identity_col_start = group_spans[0][1]
    write_group("Volume", VOLUME_COLS)
    write_group("Size", SIZE_COLS)
    write_group("Activity", ACTIVITY_COLS)
    severity_cols = {}
    for sev in SEVERITIES:
        start = col
        for sub in SEV_SUBCOLS:
            cname = f"{sev[:3]} {sub}"
            font, fill = theme.severity_fill_font(sev.lower())
            c = ws.cell(row=detail_header_row, column=col, value=cname)
            c.font, c.fill = font, fill
            severity_cols[col] = sev.lower()
            col += 1
        group_spans.append((sev, start, col - 1))

    for label, start, end in group_spans:
        cell = ws.cell(row=detail_group_row, column=start, value=label)
        theme.style_group_cell(cell, horizontal="center")
        if end > start:
            ws.merge_cells(start_row=detail_group_row, start_column=start, end_row=detail_group_row, end_column=end)

    last_col = col - 1
    ws.freeze_panes = ws.cell(row=detail_header_row + 1, column=identity_col_start + len(identity_cols))
    ws.auto_filter.ref = f"{get_column_letter(2)}{detail_header_row}:{get_column_letter(last_col)}{detail_header_row}"
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["B"].width = 34

    col_info.update(detail_group_row=detail_group_row, detail_header_row=detail_header_row,
                     detail_data_start=detail_header_row + 1, identity_col_start=identity_col_start,
                     identity_cols=identity_cols, severity_cols=severity_cols, last_col=last_col)
    col_info["identity"] = (identity_col_start,)

    # ---- Overall Distribution (stats table + 2 histograms), positioned past the
    # detail table's rightmost column so it never collides regardless of row count ----
    dist_col = last_col + 3

    theme.style_group_cell(ws.cell(row=4, column=dist_col, value=f"Overall Distribution — {sheet_name}"))
    ws.merge_cells(start_row=4, start_column=dist_col, end_row=4, end_column=dist_col + len(DISTRIBUTION_STAT_COLS) - 1)
    for i, h in enumerate(DISTRIBUTION_STAT_COLS):
        theme.style_header_cell(ws.cell(row=5, column=dist_col + i, value=h))
    dist_stats_header_row = 5
    dist_stats_data_start = 6
    dist_stats_data_end = dist_stats_data_start + len(DISTRIBUTION_METRICS) - 1

    hist_row = dist_stats_data_end + 3

    def build_histogram_block(title, hist_row):
        theme.style_group_cell(ws.cell(row=hist_row, column=dist_col, value=title))
        ws.merge_cells(start_row=hist_row, start_column=dist_col, end_row=hist_row, end_column=dist_col + 1)
        theme.style_header_cell(ws.cell(row=hist_row + 1, column=dist_col, value="Range"))
        theme.style_header_cell(ws.cell(row=hist_row + 1, column=dist_col + 1, value="Count"))
        header_row = hist_row + 1
        data_start = header_row + 1
        data_end = data_start + HISTOGRAM_BIN_COUNT - 1

        chart = BarChart()
        chart.type = "col"
        chart.title = title
        cats = Reference(ws, min_col=dist_col, min_row=data_start, max_row=data_end)
        vals = Reference(ws, min_col=dist_col + 1, min_row=header_row, max_row=data_end)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        theme.color_chart_series(chart)
        _assign_primary_axes(chart)
        chart.height, chart.width = 8, 16
        ws.add_chart(chart, f"{get_column_letter(dist_col + 3)}{hist_row}")

        return {"header_row": header_row, "data_start": data_start, "data_end": data_end}

    scans_hist = build_histogram_block(f"{sheet_name}: Scans Distribution", hist_row)
    loc_hist_row = scans_hist["data_end"] + 4
    loc_hist = build_histogram_block(f"{sheet_name}: Total LOC Distribution", loc_hist_row)

    ws.column_dimensions[get_column_letter(dist_col)].width = 16

    col_info.update(
        dist_col_start=dist_col, dist_stats_header_row=dist_stats_header_row,
        dist_stats_data_start=dist_stats_data_start, dist_stats_data_end=dist_stats_data_end,
        dist_scans_hist_header_row=scans_hist["header_row"], dist_scans_hist_data_start=scans_hist["data_start"],
        dist_loc_hist_header_row=loc_hist["header_row"], dist_loc_hist_data_start=loc_hist["data_start"],
    )
    return col_info
