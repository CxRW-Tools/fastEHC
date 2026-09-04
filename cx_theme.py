"""Checkmarx 2026 brand theme constants and styling helpers for fastEHC's Excel output.

Applying the theme at workbook-construction time (rather than patching an existing
.xlsx after the fact) means every chart series gets its color set directly at
creation -- there is no combo-chart-recoloring problem to work around.
"""
import os

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint

# ---- CX 2026 brand palette ----
QUANTUM_VIOLET = "6B34FD"
CORE_MIDNIGHT = "140921"
CLOUD_WHITE = "FCF9FE"
NEURAL_MAGENTA = "A822BF"
SIGNAL_ORANGE = "F25929"
VELOCITY_BLUE = "006BD5"

# 20/40/60% tints of Quantum Violet over white, for chart series overflow / sub-bands
VIOLET_TINT_20 = "E4D9FE"
VIOLET_TINT_40 = "C9B3FD"
VIOLET_TINT_60 = "AE8CFC"
# A subtler ~8% tint, distinct from the 20% header tint, used only for zebra striping
VIOLET_TINT_08 = "F3EEFE"

# Brand guidelines specify these exact severity colors
SEVERITY_COLORS = {
    "critical": "D91A3C",
    "high": SIGNAL_ORANGE,
    "medium": "E89000",
    "low": VELOCITY_BLUE,
    "info": "6B7280",
    "informational": "6B7280",
}

FONT_NAME = "Aptos"

GROUP_FONT = Font(name=FONT_NAME, color=CLOUD_WHITE, bold=True, size=11)
GROUP_FILL = PatternFill(start_color=QUANTUM_VIOLET, end_color=QUANTUM_VIOLET, fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, color=CORE_MIDNIGHT, bold=True, size=10)
HEADER_FILL = PatternFill(start_color=VIOLET_TINT_20, end_color=VIOLET_TINT_20, fill_type="solid")
BODY_FONT = Font(name=FONT_NAME, color=CORE_MIDNIGHT, size=10)
BODY_FILL = PatternFill(start_color=CLOUD_WHITE, end_color=CLOUD_WHITE, fill_type="solid")
# Every other data row gets a faint violet tint instead of solid Cloud White, for
# the zebra-striped readability the old template had.
BODY_FILL_ALT = PatternFill(start_color=VIOLET_TINT_08, end_color=VIOLET_TINT_08, fill_type="solid")
TITLE_FONT = Font(name=FONT_NAME, color=QUANTUM_VIOLET, bold=True, size=13)
NOTE_FONT = Font(name=FONT_NAME, italic=True, color="808080", size=9)

# No magenta -- Quantum Violet, Velocity Blue, Signal Orange, then violet tints
CHART_PALETTE = [QUANTUM_VIOLET, VELOCITY_BLUE, SIGNAL_ORANGE,
                  VIOLET_TINT_60, VIOLET_TINT_40, VIOLET_TINT_20]

THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

# Number formats -- the old template baked these into its cells; since the workbook is
# now built from scratch, every numeric cell needs its format set explicitly or Excel
# falls back to General (raw decimals for percentages, fractional days for durations).
FMT_INT = '#,##0'
FMT_DECIMAL2 = '#,##0.00'
FMT_PCT = '0.0%'
FMT_DATE = 'm/d/yyyy'
FMT_DURATION = '[h]:mm:ss'

_ASSET_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
LOGO_PATH = os.path.join(_ASSET_DIR, "Logo_-_Quantum_Violet.png")


def severity_fill_font(severity_key):
    """Return (font, fill) for a severity label/header cell."""
    sev_hex = SEVERITY_COLORS[severity_key]
    return (
        Font(name=FONT_NAME, color=CLOUD_WHITE, bold=True, size=10),
        PatternFill(start_color=sev_hex, end_color=sev_hex, fill_type="solid"),
    )


def style_group_cell(cell, horizontal="left"):
    cell.font = GROUP_FONT
    cell.fill = GROUP_FILL
    cell.alignment = Alignment(horizontal=horizontal, vertical="center")


def style_header_cell(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL


def style_body_cell(cell, alt=False):
    cell.font = BODY_FONT
    cell.fill = BODY_FILL_ALT if alt else BODY_FILL


def style_note_cell(cell):
    cell.font = NOTE_FONT


def prestyle_body_range(ws, min_row, max_row, min_col, max_col):
    """Pre-format a range of (possibly still-empty) cells as body cells, so that
    values poked in later via `cell.value = x` inherit the theme -- openpyxl only
    ever overwrites the value, never the existing style, on a `ws.cell(...)` call
    that supplies just a value onto an already-styled cell."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            style_body_cell(ws.cell(row=row, column=col))


def add_corner_logo(ws, target_height_px=28):
    if not os.path.isfile(LOGO_PATH):
        return
    ws.row_dimensions[1].height = 34
    img = XLImage(LOGO_PATH)
    scale = target_height_px / img.height
    img.height = target_height_px
    img.width = int(img.width * scale)
    ws.add_image(img, "A1")


def set_series_color(series, hex_color):
    """Set both fill (bar/pie/area) and line stroke (line/scatter) color -- a
    line-type series ignores solidFill entirely, so both must be set explicitly."""
    gp = series.graphicalProperties or GraphicalProperties()
    gp.solidFill = hex_color
    if gp.line is None:
        gp.line = LineProperties(solidFill=hex_color)
    else:
        gp.line.solidFill = hex_color
    series.graphicalProperties = gp


def color_chart_series(chart, start_index=0):
    """Cycle the brand palette across a chart's series, starting at start_index
    (useful for combo charts where a second sub-chart's series should continue
    the color sequence rather than restarting it)."""
    for i, series in enumerate(chart.series, start=start_index):
        set_series_color(series, CHART_PALETTE[i % len(CHART_PALETTE)])


def color_severity_pie(chart, order=("critical", "high", "medium", "low")):
    """Color each pie slice using the exact brand severity table."""
    for series in chart.series:
        dpts = []
        for idx, sev in enumerate(order):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties = GraphicalProperties(solidFill=SEVERITY_COLORS[sev])
            dpts.append(dp)
        series.data_points = dpts
